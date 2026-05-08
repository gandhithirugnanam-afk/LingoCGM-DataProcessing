"""
cgm_to_xls.py
=============
Converts a single person's CGM CSV data into a formatted Excel workbook that
mirrors all metrics calculated in LingoCGM_Dashboard.html.

Usage
-----
    python cgm_to_xls.py input.csv [output.xlsx]

Input CSV columns (auto-detected, case-insensitive):
    - A datetime column (header contains "time" or "date")
    - A glucose column in mg/dL (header contains "measure", "mg", or "glucose";
      or any column whose values are all in 20–700)

Output sheets
-------------
  1. Raw Data       – cleaned & sorted input readings
  2. Summary        – key scalar metrics (avg, SD, CV, GMI, eA1c, MAGE, TIR tiers,
                      fasting avg, nocturnal avg, AUC summary)
  3. AUC            – trapezoidal AUC: total / hyper (>180) / hypo (<70), daily-normalised
                      rates, and per-hour breakdown showing where burden concentrates
  4. Daily          – per-day: reading count, avg glucose, TIR%, TBR%, TAR%
  5. AGP            – hourly 10th/25th/median/75th/90th percentiles
  6. Time Blocks    – 3-hour window averages (8 windows)
                      fasting avg, nocturnal avg, best/worst days)
  3. Daily          – per-day: reading count, avg glucose, TIR%, TBR%, TAR%, AUC
  4. AGP            – hourly 10th/25th/median/75th/90th percentiles
  5. Time Blocks    – 3-hour window averages (8 windows)
"""

from __future__ import annotations

import math
import sys
import csv
import os
from datetime import datetime, date, timedelta
from collections import defaultdict
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─── Colour palette (matches dashboard CSS tokens) ────────────────────────────
C_ACCENT  = "1D5FDB"   # blue
C_GREEN   = "059669"
C_RED     = "DC2626"
C_AMBER   = "B45309"
C_PURPLE  = "7C3AED"
C_BG      = "F6F8FB"
C_SURFACE = "FFFFFF"
C_MUTED   = "7A8FAB"
C_TEXT    = "0E1624"
C_TEXT2   = "334155"
C_BORDER  = "DDE3ED"

# TIR tier colours (5-tier consensus)
C_TBR2 = "7F1D1D"   # very low <54
C_TBR1 = "DC2626"   # low 54–69
C_TIR  = "059669"   # in-range 70–180
C_TAR1 = "B45309"   # high 181–250
C_TAR2 = "78350F"   # very high >250

# Intervals longer than this (hours) are excluded from all AUC calculations
AUC_GAP_THRESHOLD_HR = 1.0


# ─── Style helpers ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_TEXT, size=11, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")


def _border_thin() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _header_cell(ws, row, col, text, bg=C_ACCENT, fg="FFFFFF",
                 bold=True, size=11, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _fill(bg)
    c.font = _font(bold=bold, color=fg, size=size)
    c.alignment = _center() if align == "center" else _left()
    c.border = _border_thin()
    return c


def _data_cell(ws, row, col, value, bold=False, color=C_TEXT,
               num_fmt=None, align="center", bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold=bold, color=color)
    c.alignment = _center() if align == "center" else _left()
    c.border = _border_thin()
    if num_fmt:
        c.number_format = num_fmt
    if bg:
        c.fill = _fill(bg)
    return c


# ─── CSV parsing ──────────────────────────────────────────────────────────────

def _detect_columns(headers: list[str]) -> tuple[str, str]:
    time_key = next(
        (h for h in headers if any(k in h.lower() for k in ("time", "date"))),
        None,
    )
    val_key = next(
        (h for h in headers
         if h != time_key and any(k in h.lower() for k in ("measure", "mg", "glucose"))),
        None,
    )
    return time_key or headers[0], val_key or headers[1]


def load_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        raise ValueError("CSV file contains no data rows.")
    headers = list(rows[0].keys())
    time_key, val_key = _detect_columns(headers)

    # Fallback: find any column whose first 20 values are all in 20–700
    if val_key not in headers or all(
        not (20 <= float(r.get(val_key, "nan") or "nan") <= 700)
        for r in rows[:20]
        if r.get(val_key, "").strip()
    ):
        for h in headers:
            if h == time_key:
                continue
            try:
                sample = [float(r[h]) for r in rows[:20] if r.get(h, "").strip()]
                if sample and all(20 <= v <= 700 for v in sample):
                    val_key = h
                    break
            except (ValueError, KeyError):
                pass

    readings: list[dict] = []
    for row in rows:
        raw = str(row.get(val_key, "")).strip()
        if not raw or raw.lower() in ("out of range", "low", "high"):
            continue
        try:
            val = float(raw)
        except ValueError:
            continue
        if not (20 <= val <= 700):
            continue
        dt_raw = str(row.get(time_key, "")).strip()
        try:
            # Try ISO / common formats
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
                        "%m/%d/%Y %H:%M", "%d/%m/%Y %H:%M",
                        "%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M:%S"):
                try:
                    dt = datetime.strptime(dt_raw, fmt)
                    break
                except ValueError:
                    continue
            else:
                dt = datetime.fromisoformat(dt_raw)
        except Exception:
            continue
        readings.append({"datetime": dt, "value": val})

    if len(readings) < 5:
        raise ValueError(
            f"Only {len(readings)} valid readings found — need ≥ 5."
        )
    readings.sort(key=lambda r: r["datetime"])
    return readings


# ─── Metric calculations (mirrors dashboard JS logic exactly) ─────────────────

def _percentile(arr: list[float], p: float) -> float:
    if not arr:
        return float("nan")
    s = sorted(arr)
    idx = (p / 100) * (len(s) - 1)
    lo, hi = int(idx), math.ceil(idx)
    return s[lo] if lo == hi else s[lo] + (s[hi] - s[lo]) * (idx - lo)


def calc_summary(readings: list[dict]) -> dict:
    vals = [r["value"] for r in readings]
    n = len(vals)
    avg = sum(vals) / n
    mn, mx = min(vals), max(vals)

    # Standard deviation (population)
    variance = sum((v - avg) ** 2 for v in vals) / n
    sd = math.sqrt(variance)

    # Coefficient of variation
    cv = (sd / avg) * 100

    # GMI — ADA/EASD: 3.31 + 0.02392 × mean_glucose
    gmi = 3.31 + 0.02392 * avg

    # eA1c — legacy formula used in dashboard
    ea1c = (avg + 46.7) / 28.7

    # MAGE — mean amplitude of glycemic excursions > 1 SD
    excursions = []
    i = 0
    while i < n:
        if vals[i] > avg + sd:
            peak = vals[i]
            while i < n and vals[i] > avg:
                peak = max(peak, vals[i])
                i += 1
            excursions.append(peak - avg)
        elif vals[i] < avg - sd:
            nadir = vals[i]
            while i < n and vals[i] < avg:
                nadir = min(nadir, vals[i])
                i += 1
            excursions.append(avg - nadir)
        else:
            i += 1
    mage = sum(excursions) / len(excursions) if excursions else sd * 1.5

    # 5-tier TIR
    tbr2 = sum(1 for v in vals if v < 54) / n * 100
    tbr1 = sum(1 for v in vals if 54 <= v < 70) / n * 100
    tir  = sum(1 for v in vals if 70 <= v <= 180) / n * 100
    tar1 = sum(1 for v in vals if 180 < v <= 250) / n * 100
    tar2 = sum(1 for v in vals if v > 250) / n * 100
    tbr  = tbr2 + tbr1   # total below range <70
    tar  = tar1 + tar2   # total above range >180

    # Hypo events (readings < 70)
    hypo_count = sum(1 for v in vals if v < 70)

    # Fasting glucose 6–10 AM
    fasting = [r["value"] for r in readings if 6 <= r["datetime"].hour < 10]
    fasting_avg = sum(fasting) / len(fasting) if fasting else None

    # Nocturnal glucose 0–4 AM
    nocturnal = [r["value"] for r in readings if 0 <= r["datetime"].hour < 4]
    nocturnal_avg = sum(nocturnal) / len(nocturnal) if nocturnal else None

    # Trend (first-half avg vs second-half avg)
    mid = n // 2
    first_avg = sum(vals[:mid]) / max(mid, 1)
    last_avg  = sum(vals[mid:]) / max(n - mid, 1)
    diff = last_avg - first_avg
    trend = "↑ Rising" if diff > 3 else ("↓ Falling" if diff < -3 else "→ Stable")

    # Days of data
    days_set = {r["datetime"].date() for r in readings}
    n_days = len(days_set)

    # Min/max timestamps
    min_r = next(r for r in readings if r["value"] == mn)
    max_r = next(r for r in readings if r["value"] == mx)

    # AUC above 70 mg/dL (trapezoidal rule, mirrors dashboard JS)
    total_auc = calc_auc(readings)
    mean_daily_auc = total_auc / n_days if n_days else 0.0

    return dict(
        n=n, avg=avg, mn=mn, mx=mx, sd=sd, cv=cv,
        gmi=gmi, ea1c=ea1c, mage=mage,
        tbr2=tbr2, tbr1=tbr1, tir=tir, tar1=tar1, tar2=tar2,
        tbr=tbr, tar=tar,
        hypo_count=hypo_count,
        fasting_avg=fasting_avg, nocturnal_avg=nocturnal_avg,
        trend=trend, n_days=n_days,
        min_dt=min_r["datetime"], max_dt=max_r["datetime"],
        date_start=min(days_set), date_end=max(days_set),
        total_auc=total_auc, mean_daily_auc=mean_daily_auc,
    )


def calc_daily(readings: list[dict]) -> list[dict]:
    by_day: dict[date, list[dict]] = defaultdict(list)
    for r in readings:
        by_day[r["datetime"].date()].append(r)
    rows = []
    for d in sorted(by_day):
        day_readings = by_day[d]
        vs = [r["value"] for r in day_readings]
        n = len(vs)
        avg = sum(vs) / n
        tir = sum(1 for v in vs if 70 <= v <= 180) / n * 100
        tbr = sum(1 for v in vs if v < 70)  / n * 100
        tar = sum(1 for v in vs if v > 180) / n * 100
        day_auc = calc_auc(day_readings)
        rows.append(dict(date=d, n=n, avg=avg, tir=tir, tbr=tbr, tar=tar, auc=day_auc))
    return rows


def calc_agp(readings: list[dict]) -> list[dict]:
    buckets: dict[int, list[float]] = defaultdict(list)
    for r in readings:
        buckets[r["datetime"].hour].append(r["value"])
    rows = []
    for h in range(24):
        vs = buckets.get(h, [])
        rows.append(dict(
            hour=h,
            label=f"{h:02d}:00",
            n=len(vs),
            p10=_percentile(vs, 10) if vs else None,
            p25=_percentile(vs, 25) if vs else None,
            p50=_percentile(vs, 50) if vs else None,
            p75=_percentile(vs, 75) if vs else None,
            p90=_percentile(vs, 90) if vs else None,
        ))
    return rows


def calc_auc(readings: list[dict], baseline: float = 70.0) -> float:
    """Trapezoidal AUC above baseline (mg/dL·h). Mirrors dashboard JS auc()."""
    pts = sorted(readings, key=lambda r: r["datetime"])
    if len(pts) < 2:
        return 0.0
    total = 0.0
    for i in range(1, len(pts)):
        dt_h = (pts[i]["datetime"] - pts[i - 1]["datetime"]).total_seconds() / 3600.0
        if dt_h > 1.0:
            continue  # skip gaps larger than 1 h (sensor dropout)
        h1 = max(0.0, pts[i - 1]["value"] - baseline)
        h2 = max(0.0, pts[i]["value"]     - baseline)
        total += (h1 + h2) / 2.0 * dt_h
    return total


def calc_time_blocks(readings: list[dict]) -> list[dict]:
    blocks = [
        ("12–3 AM",  0,  3),
        ("3–6 AM",   3,  6),
        ("6–9 AM",   6,  9),
        ("9–12 PM",  9, 12),
        ("12–3 PM", 12, 15),
        ("3–6 PM",  15, 18),
        ("6–9 PM",  18, 21),
        ("9–12 AM", 21, 24),
    ]
    rows = []
    for label, h_start, h_end in blocks:
        vs = [r["value"] for r in readings
              if h_start <= r["datetime"].hour < h_end]
        avg = sum(vs) / len(vs) if vs else None
        rows.append(dict(label=label, n=len(vs), avg=avg))
    return rows


def calc_auc(readings: list[dict]) -> dict:
    """
    Trapezoidal AUC metrics.  For each consecutive pair of readings separated
    by ≤ AUC_GAP_THRESHOLD_HR hours the contribution is:
        area = (g1 + g2) / 2 × Δt_hours
    Hyper/hypo variants use max(0, g − threshold) before averaging.
    """
    total_auc = hyper_auc = hypo_auc = valid_hours = 0.0

    for i in range(1, len(readings)):
        dt_hr = (readings[i]["datetime"] - readings[i-1]["datetime"]).total_seconds() / 3600
        if dt_hr > AUC_GAP_THRESHOLD_HR:
            continue
        g1, g2 = readings[i-1]["value"], readings[i]["value"]
        total_auc  += (g1 + g2) / 2 * dt_hr
        hyper_auc  += (max(0.0, g1 - 180) + max(0.0, g2 - 180)) / 2 * dt_hr
        hypo_auc   += (max(0.0, 70 - g1)  + max(0.0, 70 - g2))  / 2 * dt_hr
        valid_hours += dt_hr

    n_days = len({r["datetime"].date() for r in readings})
    return dict(
        total_auc=total_auc,
        hyper_auc=hyper_auc,
        hypo_auc=hypo_auc,
        daily_total=total_auc / n_days if n_days else 0.0,
        daily_hyper=hyper_auc / n_days if n_days else 0.0,
        daily_hypo=hypo_auc  / n_days if n_days else 0.0,
        valid_hours=valid_hours,
        n_days=n_days,
    )


def calc_auc_hourly(readings: list[dict]) -> list[dict]:
    """Per-hour trapezoidal AUC breakdown (interval assigned to the later reading's hour)."""
    buckets: dict[int, dict] = {
        h: {"total": 0.0, "hyper": 0.0, "hypo": 0.0, "intervals": 0}
        for h in range(24)
    }
    for i in range(1, len(readings)):
        dt_hr = (readings[i]["datetime"] - readings[i-1]["datetime"]).total_seconds() / 3600
        if dt_hr > AUC_GAP_THRESHOLD_HR:
            continue
        g1, g2 = readings[i-1]["value"], readings[i]["value"]
        h = readings[i]["datetime"].hour
        buckets[h]["total"]     += (g1 + g2) / 2 * dt_hr
        buckets[h]["hyper"]     += (max(0.0, g1 - 180) + max(0.0, g2 - 180)) / 2 * dt_hr
        buckets[h]["hypo"]      += (max(0.0, 70 - g1)  + max(0.0, 70 - g2))  / 2 * dt_hr
        buckets[h]["intervals"] += 1

    return [
        dict(hour=h, label=f"{h:02d}:00", **buckets[h])
        for h in range(24)
    ]


# ─── Sheet builders ───────────────────────────────────────────────────────────

def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def build_raw_sheet(wb: openpyxl.Workbook, readings: list[dict]):
    ws = wb.create_sheet("Raw Data")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["#", "DateTime", "Glucose (mg/dL)", "Zone"]
    widths   = [6, 22, 18, 12]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws, 1, col, h)
        _set_col_width(ws, col, w)

    for i, r in enumerate(readings, 1):
        v = r["value"]
        zone_text  = "Low"       if v < 70  else ("High" if v > 180 else "In Range")
        zone_color = C_RED       if v < 70  else (C_AMBER if v > 180 else C_GREEN)
        bg = "FEF2F2" if v < 70 else ("FFFBEB" if v > 180 else "ECFDF5")

        _data_cell(ws, i+1, 1, i, align="center")
        _data_cell(ws, i+1, 2, r["datetime"], num_fmt="YYYY-MM-DD HH:MM:SS")
        _data_cell(ws, i+1, 3, v, num_fmt="0.0", bold=(v < 54 or v > 250))
        _data_cell(ws, i+1, 4, zone_text, color=zone_color, bg=bg, bold=True)

    ws.row_dimensions[1].height = 24


def build_summary_sheet(wb: openpyxl.Workbook, s: dict, auc: dict, subject_name: str):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    def section(row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.font = _font(bold=True, color="FFFFFF", size=12)
        c.fill = _fill(C_ACCENT)
        c.alignment = _left()
        c.border = _border_thin()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    def row_metric(row, label, value, unit="", note="", color=C_TEXT, bold_val=False):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font  = _font(color=C_TEXT2)
        lc.alignment = _left()
        lc.border = _border_thin()

        vc = ws.cell(row=row, column=2, value=value)
        vc.font = _font(bold=bold_val, color=color, size=12)
        vc.alignment = _center()
        vc.border = _border_thin()
        if isinstance(value, float):
            vc.number_format = "0.0"

        uc = ws.cell(row=row, column=3, value=unit)
        uc.font = _font(color=C_MUTED, size=10)
        uc.alignment = _center()
        uc.border = _border_thin()

        nc = ws.cell(row=row, column=4, value=note)
        nc.font = _font(color=C_MUTED, size=10, italic=True)
        nc.alignment = _left()
        nc.border = _border_thin()

    _set_col_width(ws, 1, 32)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 14)
    _set_col_width(ws, 4, 42)

    # Title block
    title = ws.cell(row=1, column=1, value=f"LingoCGM — CGM Summary Report")
    title.font = _font(bold=True, size=16, color=C_ACCENT)
    title.alignment = _left()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    sub = ws.cell(row=2, column=1, value=f"Subject: {subject_name}")
    sub.font = _font(size=12, color=C_TEXT2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    date_range = f"{s['date_start']} – {s['date_end']}  ({s['n_days']} days, {s['n']:,} readings)"
    dr = ws.cell(row=3, column=1, value=date_range)
    dr.font = _font(size=10, color=C_MUTED, italic=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    ws.row_dimensions[3].height = 18

    row = 5

    # ── BASIC STATISTICS ──────────────────────────────────────────────────────
    section(row, "BASIC STATISTICS"); row += 1
    row_metric(row, "Total Readings",      s["n"],         "",       ""); row += 1
    row_metric(row, "Days of Data",        s["n_days"],    "days",   ""); row += 1
    row_metric(row, "Average Glucose",     round(s["avg"],1), "mg/dL", "", bold_val=True); row += 1
    row_metric(row, "Minimum Glucose",     s["mn"],        "mg/dL",  f"at {s['min_dt'].strftime('%Y-%m-%d %H:%M')}"); row += 1
    row_metric(row, "Maximum Glucose",     s["mx"],        "mg/dL",  f"at {s['max_dt'].strftime('%Y-%m-%d %H:%M')}"); row += 1
    row_metric(row, "Glucose Trend",       s["trend"],     "",       "Compares first-half avg vs second-half avg"); row += 1
    row_metric(row, "Hypo Events (<70)",   s["hypo_count"], "readings", "Readings below 70 mg/dL",
               color=C_RED if s["hypo_count"] > 0 else C_GREEN,
               bold_val=s["hypo_count"] > 0); row += 2

    # ── GLYCEMIC VARIABILITY ─────────────────────────────────────────────────
    section(row, "GLYCEMIC VARIABILITY"); row += 1
    row_metric(row, "Standard Deviation (SD)", round(s["sd"],1), "mg/dL",
               "Population SD of all readings"); row += 1

    cv_note = "✓ Stable (<36%)" if s["cv"] < 36 else ("⚠ Moderate (<50%)" if s["cv"] < 50 else "✕ High variability (≥50%)")
    cv_color = C_GREEN if s["cv"] < 36 else (C_AMBER if s["cv"] < 50 else C_RED)
    row_metric(row, "Coefficient of Variation (CV)", round(s["cv"],1), "%",
               cv_note, color=cv_color, bold_val=True); row += 1

    row_metric(row, "MAGE", round(s["mage"],1), "mg/dL",
               "Mean Amplitude of Glycemic Excursions (>1 SD from mean)"); row += 2

    # ── ESTIMATED HBA1C ───────────────────────────────────────────────────────
    section(row, "ESTIMATED HbA1c INDICES"); row += 1
    gmi_color = C_RED if s["gmi"] >= 7 else (C_AMBER if s["gmi"] >= 5.7 else C_GREEN)
    gmi_note = "≥7.0%: Diabetic range" if s["gmi"] >= 7 else ("5.7–6.9%: Pre-diabetic" if s["gmi"] >= 5.7 else "<5.7%: Normal")
    row_metric(row, "GMI (Glucose Mgmt Indicator)", round(s["gmi"],2), "%",
               f"3.31 + 0.02392 × {s['avg']:.1f}  —  {gmi_note}",
               color=gmi_color, bold_val=True); row += 1

    ea1c_color = C_RED if s["ea1c"] >= 6.5 else (C_AMBER if s["ea1c"] >= 5.7 else C_GREEN)
    ea1c_note = "≥6.5%: Diabetic range" if s["ea1c"] >= 6.5 else ("5.7–6.4%: Pre-diabetic" if s["ea1c"] >= 5.7 else "<5.7%: Normal")
    row_metric(row, "eA1c (legacy formula)", round(s["ea1c"],2), "%",
               f"(avg + 46.7) / 28.7  —  {ea1c_note}",
               color=ea1c_color, bold_val=True); row += 2

    # ── AUC ───────────────────────────────────────────────────────────────────
    section(row, "GLYCEMIC EXPOSURE — AREA UNDER THE CURVE (AUC > 70 mg/dL)"); row += 1
    row_metric(row, "Total AUC (all readings)", round(s["total_auc"], 1), "mg/dL·h",
               "Trapezoidal rule above 70 mg/dL baseline — Σ [(G₁−70 + G₂−70)/2] × Δt",
               bold_val=True); row += 1
    row_metric(row, "Mean Daily AUC", round(s["mean_daily_auc"], 1), "mg/dL·h/day",
               "Total AUC ÷ number of days  —  lower is better"); row += 2

    # ── 5-TIER TIR ────────────────────────────────────────────────────────────
    section(row, "5-TIER TIME IN RANGE (ADA/EASD Consensus Targets — T2D)"); row += 1

    tier_rows = [
        ("TBR Level 2  (<54 mg/dL)",    s["tbr2"], "%", "Target <1%",  C_TBR2,  1,  True),
        ("TBR Level 1  (54–69 mg/dL)",  s["tbr1"], "%", "Target <4%",  C_TBR1,  4,  True),
        ("TIR          (70–180 mg/dL)", s["tir"],  "%", "Target >70%", C_TIR,  70, False),
        ("TAR Level 1  (181–250 mg/dL)",s["tar1"], "%", "Target <25%", C_TAR1, 25,  True),
        ("TAR Level 2  (>250 mg/dL)",   s["tar2"], "%", "Target <5%",  C_TAR2,  5,  True),
    ]
    for label, val, unit, target_note, tier_color, target_val, lower_better in tier_rows:
        met = (val <= target_val) if lower_better else (val >= target_val)
        status = "✓ Met" if met else "✗ Not met"
        status_color = C_GREEN if met else C_RED
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(color=C_TEXT2); lc.alignment = _left(); lc.border = _border_thin()
        lc.fill = _fill("F0F4F9")

        vc = ws.cell(row=row, column=2, value=round(val, 1))
        vc.font = _font(bold=True, color=tier_color, size=12)
        vc.number_format = "0.0"; vc.alignment = _center(); vc.border = _border_thin()
        vc.fill = _fill("F0F4F9")

        uc = ws.cell(row=row, column=3, value=unit)
        uc.font = _font(color=C_MUTED, size=10); uc.alignment = _center(); uc.border = _border_thin()

        nc = ws.cell(row=row, column=4, value=f"{target_note}  —  {status}")
        nc.font = _font(color=status_color, size=10, italic=True, bold=met)
        nc.alignment = _left(); nc.border = _border_thin()
        row += 1
    row += 1

    # ── TIME-OF-DAY AVERAGES ──────────────────────────────────────────────────
    section(row, "TIME-OF-DAY AVERAGES"); row += 1
    for label, val in [
        ("Fasting Average (6–10 AM)",   s["fasting_avg"]),
        ("Nocturnal Average (0–4 AM)",  s["nocturnal_avg"]),
    ]:
        display = round(val, 1) if val is not None else "—"
        note = "mg/dL" if val is not None else "No readings in this window"
        row_metric(row, label, display, "mg/dL" if val is not None else "", note); row += 1
    row += 1

    # ── AREA UNDER THE CURVE ──────────────────────────────────────────────────
    section(row, "AREA UNDER THE CURVE  (Trapezoidal — gaps >1 hr excluded)"); row += 1
    row_metric(row, "Total AUC",
               round(auc["total_auc"], 1), "mg/dL·hr",
               f"Glucose-time integral over {auc['valid_hours']:.1f} valid hours "
               f"({auc['n_days']} days)"); row += 1
    row_metric(row, "Hyperglycemic AUC  (>180 mg/dL)",
               round(auc["hyper_auc"], 1), "mg/dL·hr",
               "Time-weighted exposure above 180. Captures frequency AND depth of hyperglycaemia.",
               color=C_AMBER if auc["hyper_auc"] > 0 else C_GREEN,
               bold_val=auc["hyper_auc"] > 0); row += 1
    row_metric(row, "Hypoglycemic AUC  (<70 mg/dL)",
               round(auc["hypo_auc"], 1), "mg/dL·hr",
               "Time-weighted deficit below 70. Reflects hypo severity, not just event count.",
               color=C_RED if auc["hypo_auc"] > 0 else C_GREEN,
               bold_val=auc["hypo_auc"] > 0); row += 1
    row_metric(row, "Daily Hyper AUC  (normalised)",
               round(auc["daily_hyper"], 2), "mg/dL·hr/day",
               "Hyperglycemic AUC ÷ days. Use this for comparing datasets of different lengths.",
               color=C_AMBER if auc["daily_hyper"] > 0 else C_GREEN); row += 1
    row_metric(row, "Daily Hypo AUC  (normalised)",
               round(auc["daily_hypo"], 2), "mg/dL·hr/day",
               "Hypoglycemic AUC ÷ days.",
               color=C_RED if auc["daily_hypo"] > 0 else C_GREEN); row += 1

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[5].height = 22


def build_daily_sheet(wb: openpyxl.Workbook, daily: list[dict]):
    ws = wb.create_sheet("Daily")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["Date", "Readings", "Avg Glucose (mg/dL)", "TIR % (70–180)", "TBR % (<70)", "TAR % (>180)", "AUC (mg/dL·h)", "Status"]
    widths   = [14,       10,          22,                   18,               16,             16,              16,               14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws, 1, col, h)
        _set_col_width(ws, col, w)

    for i, d in enumerate(daily, 2):
        tir_color = C_GREEN if d["tir"] >= 70 else (C_AMBER if d["tir"] >= 50 else C_RED)
        tbr_color = C_GREEN if d["tbr"] <= 4 else C_RED
        tar_color = C_GREEN if d["tar"] <= 25 else C_AMBER
        status    = "Good" if d["tir"] >= 70 and d["tbr"] <= 4 and d["tar"] <= 25 else (
                    "Low glucose" if d["tbr"] > 4 else "High glucose")
        status_color = C_GREEN if status == "Good" else C_RED if "Low" in status else C_AMBER

        _data_cell(ws, i, 1, d["date"],           num_fmt="YYYY-MM-DD", align="center")
        _data_cell(ws, i, 2, d["n"],               align="center")
        _data_cell(ws, i, 3, round(d["avg"], 1),   num_fmt="0.0", align="center")
        _data_cell(ws, i, 4, round(d["tir"], 1),   num_fmt="0.0", color=tir_color, bold=True)
        _data_cell(ws, i, 5, round(d["tbr"], 1),   num_fmt="0.0", color=tbr_color, bold=True)
        _data_cell(ws, i, 6, round(d["tar"], 1),   num_fmt="0.0", color=tar_color, bold=True)
        _data_cell(ws, i, 7, round(d["auc"], 1),   num_fmt="0.0", align="center")
        _data_cell(ws, i, 8, status, color=status_color, bold=True)

    ws.row_dimensions[1].height = 24


def build_agp_sheet(wb: openpyxl.Workbook, agp: list[dict]):
    ws = wb.create_sheet("AGP")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    title = ws.cell(row=1, column=1, value="Ambulatory Glucose Profile (AGP) — Hourly Percentiles")
    title.font = _font(bold=True, size=13, color=C_ACCENT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title.alignment = _left()
    ws.row_dimensions[1].height = 28

    sub = ws.cell(row=2, column=1,
                  value="Median glucose and interquartile range (25th–75th pct) by hour of day. "
                        "Dashed target range: 70–180 mg/dL.")
    sub.font = _font(size=10, color=C_MUTED, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    sub.alignment = _left()

    headers = ["Hour", "Readings", "10th pct", "25th pct (IQR low)",
               "50th pct (Median)", "75th pct (IQR high)", "90th pct", "Zone"]
    widths   = [10,       10,         12,          20,                   20,                  20,                  12,      14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws, 3, col, h)
        _set_col_width(ws, col, w)

    for i, h in enumerate(agp, 4):
        med = h["p50"]
        if med is None:
            for col in range(1, 9):
                _data_cell(ws, i, col, "—" if col != 1 else h["label"])
            continue
        zone_color = C_RED if med < 70 else (C_AMBER if med > 180 else C_GREEN)
        zone_text  = "Low" if med < 70 else ("High" if med > 180 else "In Range")

        def fmt(v): return round(v, 1) if v is not None else "—"

        _data_cell(ws, i, 1, h["label"])
        _data_cell(ws, i, 2, h["n"])
        _data_cell(ws, i, 3, fmt(h["p10"]), num_fmt="0.0")
        _data_cell(ws, i, 4, fmt(h["p25"]), num_fmt="0.0")
        _data_cell(ws, i, 5, fmt(h["p50"]), num_fmt="0.0", bold=True, color=C_ACCENT)
        _data_cell(ws, i, 6, fmt(h["p75"]), num_fmt="0.0")
        _data_cell(ws, i, 7, fmt(h["p90"]), num_fmt="0.0")
        _data_cell(ws, i, 8, zone_text, color=zone_color, bold=True)

    ws.row_dimensions[3].height = 24


def build_timeblock_sheet(wb: openpyxl.Workbook, blocks: list[dict]):
    ws = wb.create_sheet("Time Blocks")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    title = ws.cell(row=1, column=1, value="Average Glucose by 3-Hour Time Block")
    title.font = _font(bold=True, size=13, color=C_ACCENT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title.alignment = _left()
    ws.row_dimensions[1].height = 28

    headers = ["Time Window", "Readings", "Avg Glucose (mg/dL)", "Zone", "Clinical Relevance"]
    widths   = [16,             10,          22,                    12,      36]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws, 2, col, h)
        _set_col_width(ws, col, w)

    relevance = {
        "12–3 AM":  "Overnight baseline — hepatic glucose output",
        "3–6 AM":   "Dawn phenomenon window",
        "6–9 AM":   "Fasting / pre-breakfast",
        "9–12 PM":  "Post-breakfast",
        "12–3 PM":  "Lunch / post-prandial",
        "3–6 PM":   "Afternoon",
        "6–9 PM":   "Dinner / post-prandial",
        "9–12 AM":  "Evening / pre-sleep",
    }

    for i, b in enumerate(blocks, 3):
        avg = b["avg"]
        if avg is None:
            _data_cell(ws, i, 1, b["label"])
            for col in range(2, 6):
                _data_cell(ws, i, col, "—")
            continue
        zone_color = C_RED if avg < 70 else (C_AMBER if avg > 180 else C_GREEN)
        zone_text  = "Low" if avg < 70 else ("High" if avg > 180 else "In Range")

        _data_cell(ws, i, 1, b["label"], align="left")
        _data_cell(ws, i, 2, b["n"])
        _data_cell(ws, i, 3, round(avg, 1), num_fmt="0.0", bold=True,
                   color=zone_color)
        _data_cell(ws, i, 4, zone_text, color=zone_color, bold=True)
        rc = ws.cell(row=i, column=5, value=relevance.get(b["label"], ""))
        rc.font = _font(color=C_MUTED, size=10, italic=True)
        rc.alignment = _left()
        rc.border = _border_thin()


def build_auc_sheet(wb: openpyxl.Workbook, auc: dict, auc_hourly: list[dict]):
    ws = wb.create_sheet("AUC")
    ws.sheet_view.showGridLines = False

    # ── Title ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    t = ws.cell(row=1, column=1, value="Area Under the Glucose-Time Curve (AUC)")
    t.font = _font(bold=True, size=14, color=C_ACCENT); t.alignment = _left()
    ws.merge_cells("A1:F1")

    ws.row_dimensions[2].height = 16
    s2 = ws.cell(row=2, column=1,
                 value=f"Trapezoidal method. Sensor gaps >1 hour excluded. "
                       f"Coverage: {auc['valid_hours']:.1f} hrs over {auc['n_days']} days.")
    s2.font = _font(size=9, color=C_MUTED, italic=True); s2.alignment = _left()
    ws.merge_cells("A2:F2")

    # ── Summary metrics table ─────────────────────────────────────────────────
    ws.row_dimensions[4].height = 22
    for col, hdr in enumerate(["Metric", "Total", "Per Day (normalised)", "Unit", "Notes"], 1):
        _header_cell(ws, 4, col, hdr, bg=C_ACCENT if col != 3 else C_PURPLE)

    summary_rows = [
        ("Total AUC",
         auc["total_auc"], auc["daily_total"], "mg/dL·hr",
         "Full glucose-time integral", C_ACCENT),
        ("Hyperglycemic AUC  (>180 mg/dL)",
         auc["hyper_auc"], auc["daily_hyper"], "mg/dL·hr",
         "Exposure above 180 — frequency × depth combined", C_TAR1),
        ("Hypoglycemic AUC  (<70 mg/dL)",
         auc["hypo_auc"],  auc["daily_hypo"],  "mg/dL·hr",
         "Deficit below 70 — severity of hypos", C_TBR1),
    ]
    for i, (label, total, daily, unit, note, clr) in enumerate(summary_rows, 5):
        ws.row_dimensions[i].height = 20
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _font(color=C_TEXT2, bold=True); lc.alignment = _left(); lc.border = _border_thin()
        _data_cell(ws, i, 2, round(total, 1),  num_fmt="0.0", bold=True, color=clr)
        _data_cell(ws, i, 3, round(daily, 2),  num_fmt="0.00", bold=True, color=C_PURPLE)
        _data_cell(ws, i, 4, unit, color=C_MUTED)
        nc = ws.cell(row=i, column=5, value=note)
        nc.font = _font(color=C_MUTED, size=10, italic=True)
        nc.alignment = _left(); nc.border = _border_thin()

    # ── Per-hour breakdown ────────────────────────────────────────────────────
    ws.row_dimensions[9].height = 22
    for col, hdr in enumerate(
            ["Hour", "Intervals", "Total AUC (mg/dL·hr)",
             "Hyper AUC (>180)", "Hypo AUC (<70)", "Hyper Zone"], 1):
        _header_cell(ws, 9, col, hdr,
                     bg=C_ACCENT if col < 3 else (C_TAR1 if col == 4 else (C_TBR1 if col == 5 else C_ACCENT)))

    max_hyper = max((h["hyper"] for h in auc_hourly), default=1) or 1

    for i, h in enumerate(auc_hourly, 10):
        ws.row_dimensions[i].height = 18
        hyper_pct = h["hyper"] / max_hyper  # relative intensity 0–1
        # gradient: white → amber
        r_val = int(255)
        g_val = int(255 - 100 * hyper_pct)
        b_val = int(255 - 200 * hyper_pct)
        hyper_bg = f"{r_val:02X}{max(g_val,0):02X}{max(b_val,0):02X}" if h["hyper"] > 0 else "FFFFFF"

        _data_cell(ws, i, 1, h["label"], bold=True, color=C_ACCENT)
        _data_cell(ws, i, 2, h["intervals"])
        _data_cell(ws, i, 3, round(h["total"], 2) if h["total"] else "—", num_fmt="0.00")
        _data_cell(ws, i, 4, round(h["hyper"], 2) if h["hyper"] else "—",
                   num_fmt="0.00", bold=h["hyper"] > 0,
                   color=C_TAR1 if h["hyper"] > 0 else C_MUTED, bg=hyper_bg)
        _data_cell(ws, i, 5, round(h["hypo"],  2) if h["hypo"]  else "—",
                   num_fmt="0.00", bold=h["hypo"] > 0,
                   color=C_TBR1 if h["hypo"] > 0 else C_MUTED)
        # Hyper zone label
        pct = h["hyper"] / (h["total"] or 1) * 100
        zone = ("None" if h["hyper"] == 0 else
                "Low"  if pct < 10 else
                "Moderate" if pct < 30 else "High")
        zone_color = C_GREEN if zone == "None" else (C_AMBER if zone != "High" else C_RED)
        _data_cell(ws, i, 6, zone, color=zone_color, bold=(zone == "High"))

    # Column widths
    for col, w in enumerate([10, 10, 24, 20, 18, 14], 1):
        _set_col_width(ws, col, w)

    # Footnote
    fn_row = 35
    ws.merge_cells(f"A{fn_row}:F{fn_row}")
    fn = ws.cell(row=fn_row, column=1,
                 value="Hyper Zone %: proportion of the hour's total AUC that lies above 180 mg/dL. "
                       "High = >30%. Colour intensity in column D scales with peak-normalised hyper AUC.")
    fn.font = _font(size=9, color=C_MUTED, italic=True); fn.alignment = _left()


# ─── Main entry point ─────────────────────────────────────────────────────────

def generate_workbook(csv_path: str, out_path: str, subject_name: str = ""):
    print(f"Loading: {csv_path}")
    readings = load_csv(csv_path)
    print(f"  {len(readings):,} valid readings loaded.")

    if not subject_name:
        subject_name = os.path.splitext(os.path.basename(csv_path))[0]

    summary      = calc_summary(readings)
    auc          = calc_auc(readings)
    auc_hourly   = calc_auc_hourly(readings)
    daily        = calc_daily(readings)
    agp          = calc_agp(readings)
    timeblocks   = calc_time_blocks(readings)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    build_summary_sheet(wb, summary, auc, subject_name)
    build_auc_sheet(wb, auc, auc_hourly)
    build_daily_sheet(wb, daily)
    build_agp_sheet(wb, agp)
    build_timeblock_sheet(wb, timeblocks)
    build_raw_sheet(wb, readings)

    wb.save(out_path)
    print(f"  Saved: {out_path}")
    _print_summary(summary, subject_name)
    _print_auc(auc)


def _print_summary(s: dict, name: str):
    print(f"\n{'─'*50}")
    print(f"  {name}")
    print(f"{'─'*50}")
    print(f"  Readings : {s['n']:,}  |  Days: {s['n_days']}")
    print(f"  Avg      : {s['avg']:.1f} mg/dL  (SD: {s['sd']:.1f}, CV: {s['cv']:.1f}%)")
    print(f"  GMI      : {s['gmi']:.2f}%  |  eA1c: {s['ea1c']:.2f}%")
    print(f"  MAGE     : {s['mage']:.1f} mg/dL")
    print(f"  AUC>70   : {s['total_auc']:.1f} mg/dL·h total  |  {s['mean_daily_auc']:.1f} mg/dL·h/day avg")
    print(f"  TIR      : {s['tir']:.1f}%  (TBR {s['tbr']:.1f}%  /  TAR {s['tar']:.1f}%)")
    print(f"  5-tier   : TBR2 {s['tbr2']:.1f}%  TBR1 {s['tbr1']:.1f}%  TIR {s['tir']:.1f}%"
          f"  TAR1 {s['tar1']:.1f}%  TAR2 {s['tar2']:.1f}%")
    if s["fasting_avg"]:
        print(f"  Fasting  : {s['fasting_avg']:.1f} mg/dL")
    if s["nocturnal_avg"]:
        print(f"  Nocturnal: {s['nocturnal_avg']:.1f} mg/dL")
    print(f"  Trend    : {s['trend']}")
    print(f"  Hypos    : {s['hypo_count']} readings <70 mg/dL")
    print(f"{'─'*50}\n")


def _print_auc(auc: dict):
    print(f"  AUC (total)   : {auc['total_auc']:>10.1f} mg/dL·hr  "
          f"({auc['daily_total']:.1f}/day)")
    print(f"  AUC (hyper)   : {auc['hyper_auc']:>10.1f} mg/dL·hr  "
          f"({auc['daily_hyper']:.2f}/day  >180 mg/dL)")
    print(f"  AUC (hypo)    : {auc['hypo_auc']:>10.1f} mg/dL·hr  "
          f"({auc['daily_hypo']:.2f}/day  <70 mg/dL)")
    print(f"  Coverage      : {auc['valid_hours']:.1f} hrs  "
          f"({auc['n_days']} days,  gaps >1 hr excluded)")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nUsage: python cgm_to_xls.py input.csv [output.xlsx] [\"Subject Name\"]")
        sys.exit(1)

    csv_path = sys.argv[1]
    if not os.path.isfile(csv_path):
        print(f"Error: file not found: {csv_path}")
        sys.exit(1)

    base = os.path.splitext(csv_path)[0]
    out_path = sys.argv[2] if len(sys.argv) >= 3 else base + "_cgm_report.xlsx"
    subject  = sys.argv[3] if len(sys.argv) >= 4 else ""

    generate_workbook(csv_path, out_path, subject)


if __name__ == "__main__":
    main()
