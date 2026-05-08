"""
create_cgm_template.py
======================
Generates LingoCGM_CGM_Report.xlsx — a self-contained Excel workbook.

The user pastes their CGM data into the INPUT sheet (columns A & B).
All other sheets auto-recalculate using Excel formulas (requires Excel 365
or Excel 2021 for the Daily and AGP sheets which use FILTER/UNIQUE/SORT).

Sheets produced
---------------
  INPUT        — paste DateTime + Glucose data here (up to 20 000 readings)
                 Cell I1: Metformin start date  |  Cell I2: Washout buffer (days, default 4)
  Summary      — scalar metrics: avg, SD, CV, LBGI, HBGI, GMI, HBS, 5-tier TIR
  Comparative  — Off Meds vs On Meds: all metrics, percentile distribution, time blocks
  Daily        — per-day: count, avg glucose, TIR%, TBR%, TAR%, rolling 14-day TIR, nocturnal hypos
  AGP          — hourly 10th / 25th / median / 75th / 90th percentiles
  Time Blocks  — 3-hour window averages with Off Meds / On Meds split
  _Calc        — hidden helper (date, hour, med status, Kovatchev risk per row; do not edit)

Run:  python create_cgm_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

NROWS   = 20_000
DSTART  = 3
DEND    = DSTART + NROWS - 1

# ── Colour palette ────────────────────────────────────────────────────────────

C = dict(
    accent  = "1D5FDB",
    green   = "059669",
    red     = "DC2626",
    amber   = "B45309",
    purple  = "7C3AED",
    bg      = "F6F8FB",
    surface = "FFFFFF",
    muted   = "7A8FAB",
    text    = "0E1624",
    text2   = "334155",
    border  = "DDE3ED",
    tbr2    = "7F1D1D",
    tbr1    = "DC2626",
    tir     = "059669",
    tar1    = "B45309",
    tar2    = "78350F",
    input_g = "ECFDF5",
    input_d = "EFF4FF",
    hdr_alt = "F0F4F9",
)

# ── Style helpers ─────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="0E1624", size=11, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def border():
    s = Side(style="thin", color="DDE3ED")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="center", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def col_w(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def hcell(ws, row, col, text, bg=None, fg="FFFFFF", bold=True, size=11,
          h_align="center", wrap=False):
    c = ws.cell(row=row, column=col, value=text)
    c.fill  = fill(bg or C["accent"])
    c.font  = font(bold=bold, color=fg, size=size)
    c.alignment = align(h_align, wrap=wrap)
    c.border = border()
    return c

def dcell(ws, row, col, value, bold=False, color=None, num_fmt=None,
          h_align="center", bg=None, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font  = font(bold=bold, color=color or C["text"], italic=italic)
    c.alignment = align(h_align)
    c.border = border()
    if num_fmt:
        c.number_format = num_fmt
    if bg:
        c.fill = fill(bg)
    return c

def merge_hdr(ws, row, c1, c2, text, bg=None, fg="FFFFFF", size=12, h_align="left"):
    hcell(ws, row, c1, text, bg=bg, fg=fg, size=size, h_align=h_align)
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

# ── Formula range shorthands ──────────────────────────────────────────────────

GLU  = f"_Calc!$B${DSTART}:$B${DEND}"
DATE = f"_Calc!$C${DSTART}:$C${DEND}"
HOUR = f"_Calc!$D${DSTART}:$D${DEND}"
MEDS = f"_Calc!$E${DSTART}:$E${DEND}"
RL        = f"_Calc!$G${DSTART}:$G${DEND}"   # Kovatchev low-risk component → LBGI
RH        = f"_Calc!$H${DSTART}:$H${DEND}"   # Kovatchev high-risk component → HBGI
TRAP_AUC   = f"_Calc!$J${DSTART}:$J${DEND}"   # Trapezoidal total AUC contributions
TRAP_HYPER = f"_Calc!$K${DSTART}:$K${DEND}"   # Trapezoidal AUC > 180 contributions
TRAP_HYPO  = f"_Calc!$L${DSTART}:$L${DEND}"   # Trapezoidal AUC < 70 contributions
N    = f"COUNT({GLU})"

MED_START_CELL   = "INPUT!$I$1"
MED_WASHOUT_CELL = "INPUT!$I$2"

# ── Formula builders ─────────────────────────────────────────────────────────

def tir_formula(lo, hi):
    return f"=IFERROR(COUNTIFS({GLU},\">=\"&{lo},{GLU},\"<=\"&{hi})/{N}*100,0)"

def countif_f(crit):
    return f"=IFERROR(COUNTIF({GLU},{crit})/{N}*100,0)"

def countifs_f(c1, c2):
    return f"=IFERROR(COUNTIFS({GLU},\">=\"&{c1},{GLU},\"<\"&{c2})/{N}*100,0)"

def avgifs_hour(h_lo, h_hi):
    return (f"=IFERROR(AVERAGEIFS({GLU},{HOUR},\">=\"&{h_lo},"
            f"{HOUR},\"<\"&{h_hi}),\"—\")")

def pctile_hour(h, p):
    return f"=IFERROR(PERCENTILE(FILTER({GLU},{HOUR}={h}),{p}),\"—\")"

# Medication-conditioned builders
def avgifs_meds(phase):
    return f'=IFERROR(AVERAGEIFS({GLU},{MEDS},"{phase}"),"—")'

def countif_meds(phase):
    return f'=IFERROR(COUNTIF({MEDS},"{phase}"),0)'

def sd_meds(phase):
    avg_expr = f'AVERAGEIFS({GLU},{MEDS},"{phase}")'
    safe_val = f'IF(ISNUMBER({GLU}),{GLU},0)'
    return (f'=IFERROR(SQRT(SUMPRODUCT(({MEDS}="{phase}")*ISNUMBER({GLU})*'
            f'({safe_val}-{avg_expr})^2)/COUNTIF({MEDS},"{phase}")),"—")')

def gmi_meds(phase):
    return f'=IFERROR(3.31+0.02392*AVERAGEIFS({GLU},{MEDS},"{phase}"),"—")'

def avgifs_hour_meds(h_lo, h_hi, phase):
    return (f'=IFERROR(AVERAGEIFS({GLU},{HOUR},">="&{h_lo},'
            f'{HOUR},"<"&{h_hi},{MEDS},"{phase}"),"—")')

def lbgi_meds(phase):
    return f'=IFERROR(AVERAGEIFS({RL},{MEDS},"{phase}"),"—")'

def hbgi_meds(phase):
    return f'=IFERROR(AVERAGEIFS({RH},{MEDS},"{phase}"),"—")'

def hbs_meds(phase):
    return (f'=IFERROR(SUMPRODUCT(({MEDS}="{phase}")*({GLU}>180)*'
            f'(IF(ISNUMBER({GLU}),{GLU},0)-180)),"—")')

def pctile_meds(p, phase):
    return f'=IFERROR(PERCENTILE(FILTER({GLU},{MEDS}="{phase}"),{p}),"—")'

def auc_sumif(trap_range, phase):
    return f'=IFERROR(SUMIF({MEDS},"{phase}",{trap_range}),"—")'

def daily_auc_meds(trap_range, phase):
    """AUC per phase normalised by unique days in that phase."""
    days = (f'SUMPRODUCT(({MEDS}="{phase}")'
            f'*IFERROR(1/COUNTIFS({MEDS},"{phase}",{DATE},{DATE}),0))')
    return f'=IFERROR(SUMIF({MEDS},"{phase}",{trap_range})/({days}),"—")'

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def build_input(wb):
    ws = wb.create_sheet("INPUT")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.tab_color = C["accent"]

    ws.row_dimensions[1].height = 36

    # ── Title banner (A1:B1) ─────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "LingoCGM — CGM Data Input"
    t.font  = font(bold=True, size=16, color="FFFFFF")
    t.fill  = fill(C["accent"])
    t.alignment = align("left")

    # ── Instructions (C1:F1) ─────────────────────────────────────────────────
    inst = ws["C1"]
    inst.value = (
        f"Paste your CGM data below (rows 3–{DEND:,}).  "
        "Column A = DateTime  |  Column B = Glucose (mg/dL).  "
        "All other sheets recalculate automatically.  "
        "Maximum 20 000 readings (~70 days at 5-min, ~208 days at 15-min).  "
        "Requires Excel 365 / Excel 2021."
    )
    inst.font  = font(size=9, color="FFFFFF", italic=True)
    inst.fill  = fill(C["accent"])
    inst.alignment = align("left", wrap=True)
    ws.merge_cells("C1:F1")

    # ── G1/G2 spacer ─────────────────────────────────────────────────────────
    for gr in [1, 2]:
        g = ws.cell(row=gr, column=7)
        g.fill = fill(C["bg"])
        g.border = border()

    # ── H1: label  I1: date input ─────────────────────────────────────────────
    h1 = ws.cell(row=1, column=8, value="Metformin Start Date:")
    h1.font = font(bold=True, size=10, color="FFFFFF")
    h1.fill = fill(C["amber"])
    h1.alignment = align("left")
    h1.border = border()

    i1 = ws.cell(row=1, column=9)
    i1.number_format = "YYYY-MM-DD"
    i1.font = font(bold=True, size=11, color=C["text"])
    i1.fill = fill("FFFBEB")   # light yellow = user input
    i1.alignment = align("center")
    i1.border = border()

    # ── Row 2 column headers ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 22
    hcell(ws, 2, 1, "DateTime",        bg=C["accent"], size=12)
    hcell(ws, 2, 2, "Glucose (mg/dL)", bg="059669",    size=12)
    hcell(ws, 2, 3, "← Paste data in columns A & B from row 3 onwards",
          bg=C["hdr_alt"], fg=C["muted"], bold=False, size=10, h_align="left")
    ws.merge_cells("C2:F2")

    # ── H2: label  I2: washout buffer ────────────────────────────────────────
    h2 = ws.cell(row=2, column=8, value="Washout Buffer (days):")
    h2.font = font(bold=True, size=10, color="FFFFFF")
    h2.fill = fill(C["amber"])
    h2.alignment = align("left")
    h2.border = border()

    i2 = ws.cell(row=2, column=9, value=4)
    i2.font = font(bold=True, size=11, color=C["text"])
    i2.fill = fill("FFFBEB")
    i2.alignment = align("center")
    i2.border = border()
    i2.number_format = "0"

    # ── Column widths ────────────────────────────────────────────────────────
    col_w(ws, 1, 22); col_w(ws, 2, 18); col_w(ws, 3, 48)
    col_w(ws, 7, 3);  col_w(ws, 8, 26); col_w(ws, 9, 20)

    # ── Data rows: format A as datetime, B as number ──────────────────────────
    for r in range(DSTART, DEND + 1):
        a = ws.cell(row=r, column=1)
        a.number_format = "YYYY-MM-DD HH:MM:SS"
        a.fill = fill("EFF4FF")
        a.alignment = align("left")
        a.border = border()

        b = ws.cell(row=r, column=2)
        b.number_format = "0.0"
        b.fill = fill("ECFDF5")
        b.alignment = align("center")
        b.border = border()

    # ── Conditional formatting: colour glucose by zone ───────────────────────
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import PatternFill as PF

    glu_range = f"B{DSTART}:B{DEND}"
    ws.conditional_formatting.add(glu_range,
        CellIsRule(operator="lessThan", formula=["70"],
                   fill=PF("solid", fgColor="FEF2F2"),
                   font=Font(color=C["red"],   bold=True, name="Calibri")))
    ws.conditional_formatting.add(glu_range,
        CellIsRule(operator="greaterThan", formula=["180"],
                   fill=PF("solid", fgColor="FFFBEB"),
                   font=Font(color=C["amber"], bold=True, name="Calibri")))
    ws.conditional_formatting.add(glu_range,
        FormulaRule(formula=[f"AND(B{DSTART}>=70,B{DSTART}<=180)"],
                    fill=PF("solid", fgColor="ECFDF5"),
                    font=Font(color=C["green"], name="Calibri")))


def build_calc(wb):
    """Hidden helper: valid glucose, date serial, hour, med status, Kovatchev risk per row."""
    ws = wb.create_sheet("_Calc")
    ws.sheet_view.showGridLines = False
    ws.sheet_state = "hidden"

    for col, hdr in enumerate(
            ["DateTime", "Glucose_Valid", "Date_Serial", "Hour",
             "Med_Status", "f_val", "rl_lbgi", "rh_hbgi",
             "dt_hr", "trap_auc", "trap_hyper", "trap_hypo"], 1):
        ws.cell(row=1, column=col, value=hdr)

    print(f"  Writing {NROWS:,} _Calc helper rows…", end="", flush=True)
    for r in range(DSTART, DEND + 1):
        ws.cell(row=r, column=1, value=f"=Input!A{r}")
        ws.cell(row=r, column=2, value=(
            f'=IF(ISNUMBER(Input!B{r}),'
            f'IF(AND(Input!B{r}>=20,Input!B{r}<=700),Input!B{r},""),"")'))
        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",INT(A{r}))')
        ws.cell(row=r, column=4, value=f'=IF(B{r}="","",HOUR(A{r}))')
        # Med_Status: Off Meds / Washout / On Meds based on single start date
        ws.cell(row=r, column=5, value=(
            f'=IF(B{r}="","",IF({MED_START_CELL}="","",IF(C{r}<(INT({MED_START_CELL})-{MED_WASHOUT_CELL}),'
            f'"Off Meds",IF(C{r}<=(INT({MED_START_CELL})+{MED_WASHOUT_CELL}),"Washout","On Meds"))))'))
        # Kovatchev risk transform: f(BG) = 1.509 * (ln(BG)^1.084 - 5.381)
        ws.cell(row=r, column=6, value=f'=IF(B{r}="","",1.509*(LN(B{r})^1.084-5.381))')
        # rl: low-risk component (f < 0 → hypoglycaemia risk)
        ws.cell(row=r, column=7, value=f'=IF(F{r}="","",IF(F{r}<0,10*F{r}^2,0))')
        # rh: high-risk component (f > 0 → hyperglycaemia risk)
        ws.cell(row=r, column=8, value=f'=IF(F{r}="","",IF(F{r}>0,10*F{r}^2,0))')
        # dt_hr: hours elapsed since previous valid reading; "" if first row or gap > 1 hr
        ws.cell(row=r, column=9, value=(
            f'=IF(B{r}="","",IF(ROW()<={DSTART},"",IF(B{r-1}="","",IF((A{r}-A{r-1})*24>1,"gap",(A{r}-A{r-1})*24))))'))
        # trap_auc: trapezoidal total AUC contribution (mg/dL·hr) for this interval
        ws.cell(row=r, column=10, value=(
            f'=IF(OR(I{r}="",I{r}="gap"),0,(B{r}+B{r-1})/2*I{r})'))
        # trap_hyper: trapezoidal AUC above 180 mg/dL
        ws.cell(row=r, column=11, value=(
            f'=IF(OR(I{r}="",I{r}="gap"),0,(MAX(0,B{r}-180)+MAX(0,B{r-1}-180))/2*I{r})'))
        # trap_hypo: trapezoidal AUC below 70 mg/dL
        ws.cell(row=r, column=12, value=(
            f'=IF(OR(I{r}="",I{r}="gap"),0,(MAX(0,70-B{r})+MAX(0,70-B{r-1}))/2*I{r})'))
        if r % 5000 == 0:
            print(".", end="", flush=True)
    print(" done")

    for i, w in enumerate([22, 14, 14, 8, 12, 14, 12, 12, 10, 12, 12, 12], 1):
        col_w(ws, i, w)


def build_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["accent"]
    col_w(ws, 1, 36); col_w(ws, 2, 18); col_w(ws, 3, 14); col_w(ws, 4, 46)

    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "LingoCGM — CGM Summary Report"
    t.font  = font(bold=True, size=16, color=C["accent"])
    t.alignment = align("left")

    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = "Edit the subject name here →"
    sub.font  = font(size=12, color=C["text2"], italic=True)
    sub.alignment = align("left")

    r = 4

    def section(title, bg=None):
        nonlocal r
        ws.row_dimensions[r].height = 22
        merge_hdr(ws, r, 1, 4, title, bg=bg or C["accent"], size=12)
        r += 1

    def metric_row(label, formula, unit="", note="", val_color=None,
                   bold_val=False, num_fmt="0.0", bg_val=None):
        nonlocal r
        ws.row_dimensions[r].height = 20
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font(color=C["text2"]); lc.alignment = align("left"); lc.border = border()
        vc = ws.cell(row=r, column=2, value=formula)
        vc.font = font(bold=bold_val, color=val_color or C["text"], size=12)
        vc.alignment = align("center"); vc.border = border()
        vc.number_format = num_fmt
        if bg_val: vc.fill = fill(bg_val)
        uc = ws.cell(row=r, column=3, value=unit)
        uc.font = font(color=C["muted"], size=10); uc.alignment = align("center"); uc.border = border()
        nc = ws.cell(row=r, column=4, value=note)
        nc.font = font(color=C["muted"], size=10, italic=True)
        nc.alignment = align("left", wrap=True); nc.border = border()
        r += 1

    def blank():
        nonlocal r
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = border()
        r += 1

    # ── OVERVIEW ─────────────────────────────────────────────────────────────
    section("OVERVIEW")
    metric_row("Subject / File",  "← enter name in cell B2", note="Edit B2 above")
    metric_row("Total Readings",  f"={N}", unit="readings", num_fmt="0", bold_val=True)
    metric_row("Days of Data",
               f"=IFERROR(SUMPRODUCT((1/COUNTIFS({DATE},{DATE},{GLU},\"<>\"))*({GLU}<>\"\"),\"\"),\"—\")",
               unit="days", num_fmt="0", note="Unique calendar days with ≥1 valid reading")
    metric_row("Date Range (start)",
               f"=IFERROR(TEXT(MIN({DATE}),\"YYYY-MM-DD\"),\"—\")",
               num_fmt="@", note="Earliest date in dataset")
    metric_row("Date Range (end)",
               f"=IFERROR(TEXT(MAX({DATE}),\"YYYY-MM-DD\"),\"—\")",
               num_fmt="@", note="Latest date in dataset")
    blank()

    # ── BASIC STATISTICS ─────────────────────────────────────────────────────
    section("BASIC STATISTICS")
    metric_row("Average Glucose",
               f"=IFERROR(AVERAGE({GLU}),\"—\")",
               unit="mg/dL", bold_val=True, note="Mean of all valid readings")
    metric_row("Minimum Glucose",
               f"=IFERROR(MIN({GLU}),\"—\")",
               unit="mg/dL", note="Lowest single reading")
    metric_row("Maximum Glucose",
               f"=IFERROR(MAX({GLU}),\"—\")",
               unit="mg/dL", note="Highest single reading")
    metric_row("Hypo Events (<70 mg/dL)",
               f"=IFERROR(COUNTIF({GLU},\"<70\"),\"—\")",
               unit="readings", num_fmt="0", val_color=C["red"],
               note="Total readings below 70 mg/dL (ADA hypoglycaemia threshold)")
    metric_row("Hyperglycemia Burden Score (HBS)",
               f'=IFERROR(SUMPRODUCT(({GLU}>180)*(IF(ISNUMBER({GLU}),{GLU},0)-180)),"—")',
               unit="mg/dL · rdgs",
               note="Sum of excess glucose above 180 mg/dL. Captures frequency AND magnitude — "
                    "a tighter treatment-effect signal than TAR% alone.")
    blank()

    # ── GLYCEMIC VARIABILITY ─────────────────────────────────────────────────
    section("GLYCEMIC VARIABILITY")
    metric_row("Standard Deviation (SD)",
               f"=IFERROR(STDEV.P({GLU}),\"—\")",
               unit="mg/dL", note="Population SD — how spread glucose values are")
    metric_row("Coefficient of Variation (CV)",
               f"=IFERROR(STDEV.P({GLU})/AVERAGE({GLU})*100,\"—\")",
               unit="%",
               note="CV = SD ÷ Mean × 100.  Target: <36% (stable).  36–50%: moderate.  >50%: high variability")
    metric_row("MAGE",
               "— (see note)", num_fmt="@",
               note="Mean Amplitude of Glycemic Excursions cannot be calculated with Excel formulas. "
                    "Use the Python script (cgm_to_xls.py) to obtain MAGE.")
    metric_row("LBGI — Low Blood Glucose Index",
               f'=IFERROR(AVERAGEIF({RL},"<>"),"—")',
               unit="score",
               note="Kovatchev FDA risk score for hypoglycaemia. "
                    "<1.1: Low risk.  1.1–2.5: Moderate.  >2.5: High.")
    metric_row("HBGI — High Blood Glucose Index",
               f'=IFERROR(AVERAGEIF({RH},"<>"),"—")',
               unit="score",
               note="Kovatchev FDA risk score for hyperglycaemia. "
                    "<4.5: Low risk.  4.5–9: Moderate.  >9: High.  Should drop on metformin.")
    blank()

    # ── ESTIMATED HBA1C ───────────────────────────────────────────────────────
    section("ESTIMATED HbA1c")
    metric_row("GMI — Glucose Management Indicator",
               f"=IFERROR(3.31+0.02392*AVERAGE({GLU}),\"—\")",
               unit="%", val_color=C["accent"], bold_val=True,
               note="ADA/EASD formula: 3.31 + 0.02392 × mean glucose (mg/dL).  "
                    "≥7.0%: Diabetic  5.7–6.9%: Pre-diabetic  <5.7%: Normal")
    metric_row("eA1c (legacy)",
               f"=IFERROR((AVERAGE({GLU})+46.7)/28.7,\"—\")",
               unit="%",
               note="Legacy estimate: (avg + 46.7) / 28.7.  GMI is preferred for CGM data.")
    blank()

    # ── 5-TIER TIR ────────────────────────────────────────────────────────────
    section("5-TIER TIME IN RANGE  (ADA/EASD Consensus Targets — T2D)")
    tier_rows = [
        ("TBR Level 2  (<54 mg/dL)",
         f"=IFERROR(COUNTIF({GLU},\"<54\")/{N}*100,0)", C["tbr2"], "Target <1%"),
        ("TBR Level 1  (54–69 mg/dL)",
         f"=IFERROR(COUNTIFS({GLU},\">=\"&54,{GLU},\"<\"&70)/{N}*100,0)", C["tbr1"], "Target <4%"),
        ("TIR  (70–180 mg/dL)",
         f"=IFERROR(COUNTIFS({GLU},\">=\"&70,{GLU},\"<=\"&180)/{N}*100,0)", C["tir"], "Target >70%"),
        ("TAR Level 1  (181–250 mg/dL)",
         f"=IFERROR(COUNTIFS({GLU},\">\"&180,{GLU},\"<=\"&250)/{N}*100,0)", C["tar1"], "Target <25%"),
        ("TAR Level 2  (>250 mg/dL)",
         f"=IFERROR(COUNTIF({GLU},\">250\")/{N}*100,0)", C["tar2"], "Target <5%"),
    ]
    for label, formula, clr, target in tier_rows:
        ws.row_dimensions[r].height = 20
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font(color=C["text2"]); lc.alignment = align("left")
        lc.border = border(); lc.fill = fill(C["hdr_alt"])
        vc = ws.cell(row=r, column=2, value=formula)
        vc.font = font(bold=True, color=clr, size=12)
        vc.number_format = "0.0"; vc.alignment = align("center")
        vc.border = border(); vc.fill = fill(C["hdr_alt"])
        uc = ws.cell(row=r, column=3, value="%")
        uc.font = font(color=C["muted"], size=10); uc.alignment = align("center"); uc.border = border()
        nc = ws.cell(row=r, column=4, value=target)
        nc.font = font(color=C["muted"], size=10, italic=True)
        nc.alignment = align("left"); nc.border = border()
        r += 1
    blank()

    # ── TIME-OF-DAY AVERAGES ──────────────────────────────────────────────────
    section("TIME-OF-DAY AVERAGES")
    metric_row("Fasting Average (6–10 AM)",
               avgifs_hour(6, 10), unit="mg/dL", num_fmt='0.0"  mg/dL"',
               note="Readings between 06:00–09:59. Key metformin efficacy marker.")
    metric_row("Nocturnal Average (0–4 AM)",
               avgifs_hour(0, 4), unit="mg/dL",
               note="Readings between 00:00–03:59. Elevated values suggest hepatic insulin resistance.")
    blank()

    # ── SIMPLIFIED TIR ────────────────────────────────────────────────────────
    section("SIMPLIFIED TIR  (3-tier)", bg=C["text2"])
    metric_row("TBR Total  (<70 mg/dL)",
               f"=IFERROR(COUNTIF({GLU},\"<70\")/{N}*100,0)",
               unit="%", val_color=C["red"], bold_val=True,
               note="All readings below target range. ADA target <4%.")
    metric_row("TIR  (70–180 mg/dL)",
               f"=IFERROR(COUNTIFS({GLU},\">=\"&70,{GLU},\"<=\"&180)/{N}*100,0)",
               unit="%", val_color=C["green"], bold_val=True,
               note="Readings within target range. ADA target >70%.")
    metric_row("TAR Total  (>180 mg/dL)",
               f"=IFERROR(COUNTIF({GLU},\">180\")/{N}*100,0)",
               unit="%", val_color=C["amber"], bold_val=True,
               note="All readings above target range. ADA target <25%.")
    blank()

    # ── AREA UNDER THE CURVE ──────────────────────────────────────────────────
    section("AREA UNDER THE CURVE  (Trapezoidal Method)", bg=C["purple"])
    # Total valid hours: sum of dt_hr where dt_hr is a number (excludes gaps)
    _dt  = f"_Calc!$I${DSTART}:$I${DEND}"
    _days = f'IFERROR(SUMPRODUCT((1/COUNTIFS({DATE},{DATE},{GLU},"<>"))*({GLU}<>"")),1)'
    metric_row("Total AUC",
               f"=IFERROR(SUM({TRAP_AUC}),\"—\")",
               unit="mg/dL·hr",
               note="Integral of the glucose-time curve over the entire dataset. "
                    "Gaps >1 hour between readings are excluded.")
    metric_row("Hyperglycemic AUC  (>180 mg/dL)",
               f"=IFERROR(SUM({TRAP_HYPER}),\"—\")",
               unit="mg/dL·hr", val_color=C["amber"],
               note="Time-weighted glucose exposure above 180 mg/dL. "
                    "Captures both frequency and depth of hyperglycaemia.")
    metric_row("Hypoglycemic AUC  (<70 mg/dL)",
               f"=IFERROR(SUM({TRAP_HYPO}),\"—\")",
               unit="mg/dL·hr", val_color=C["red"],
               note="Time-weighted glucose deficit below 70 mg/dL. "
                    "Reflects hypoglycaemia severity, not just event count.")
    metric_row("Daily Hyper AUC  (per calendar day)",
               f'=IFERROR(SUM({TRAP_HYPER})/({_days}),"—")',
               unit="mg/dL·hr/day", val_color=C["amber"],
               note="Hyperglycemic AUC normalised by unique days of data. "
                    "Enables fair comparison across datasets of different lengths.")
    metric_row("Daily Hypo AUC  (per calendar day)",
               f'=IFERROR(SUM({TRAP_HYPO})/({_days}),"—")',
               unit="mg/dL·hr/day", val_color=C["red"],
               note="Hypoglycemic AUC normalised by unique days of data.")


def build_comparative_summary(wb):
    """Side-by-side Off Meds vs On Meds analysis (Suggestions 2 & 3)."""
    ws = wb.create_sheet("Comparative")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["purple"]

    col_w(ws, 1, 36); col_w(ws, 2, 18); col_w(ws, 3, 18)
    col_w(ws, 4, 16); col_w(ws, 5, 14); col_w(ws, 6, 44)

    r = 1

    # ── Title ────────────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 32
    ws.merge_cells(f"A{r}:F{r}")
    t = ws.cell(row=r, column=1, value="LingoCGM — Metformin Comparative Analysis")
    t.font = font(bold=True, size=16, color=C["accent"]); t.alignment = align("left")
    r += 1

    ws.row_dimensions[r].height = 16
    ws.merge_cells(f"A{r}:F{r}")
    s = ws.cell(row=r, column=1,
                value="Side-by-side CGM metrics for Off Meds vs On Meds periods. "
                      "Enter your metformin start date in INPUT sheet cell I1. "
                      "Negative Δ = lower on meds (favourable for most metrics).")
    s.font = font(size=9, color=C["muted"], italic=True); s.alignment = align("left")
    r += 1
    r += 1  # spacer

    # ── Local helpers ─────────────────────────────────────────────────────────
    def section(title, bg=None):
        nonlocal r
        ws.row_dimensions[r].height = 22
        hcell(ws, r, 1, title, bg=bg or C["accent"], size=12, h_align="left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    def col_hdrs(h2="Off Meds", h3="On Meds"):
        nonlocal r
        ws.row_dimensions[r].height = 20
        hcell(ws, r, 1, "Metric",   bg=C["text2"])
        hcell(ws, r, 2, h2,         bg=C["tbr1"])
        hcell(ws, r, 3, h3,         bg=C["tir"])
        hcell(ws, r, 4, "Δ Change", bg=C["purple"])
        hcell(ws, r, 5, "Unit",     bg=C["accent"])
        hcell(ws, r, 6, "Notes",    bg=C["accent"])
        r += 1

    def mrow(label, f_off, f_on, unit="", note="", num_fmt="0.0",
             off_color=None, on_color=None, no_delta=False):
        nonlocal r
        ws.row_dimensions[r].height = 20

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font(color=C["text2"]); lc.alignment = align("left"); lc.border = border()

        oc = ws.cell(row=r, column=2, value=f_off)
        oc.font = font(bold=True, color=off_color or C["tbr1"], size=11)
        oc.alignment = align("center"); oc.border = border(); oc.number_format = num_fmt

        nc = ws.cell(row=r, column=3, value=f_on)
        nc.font = font(bold=True, color=on_color or C["tir"], size=11)
        nc.alignment = align("center"); nc.border = border(); nc.number_format = num_fmt

        dv = ("—" if no_delta else
              f'=IFERROR(IF(ISNUMBER(B{r})*ISNUMBER(C{r}),C{r}-B{r},"—"),"—")')
        dc = ws.cell(row=r, column=4, value=dv)
        dc.font = font(bold=True, color=C["purple"], size=11)
        dc.alignment = align("center"); dc.border = border(); dc.number_format = num_fmt

        uc = ws.cell(row=r, column=5, value=unit)
        uc.font = font(color=C["muted"], size=10); uc.alignment = align("center"); uc.border = border()

        ntc = ws.cell(row=r, column=6, value=note)
        ntc.font = font(color=C["muted"], size=10, italic=True)
        ntc.alignment = align("left", wrap=True); ntc.border = border()
        r += 1

    def blank():
        nonlocal r
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = border()
        r += 1

    # ── CONFIGURATION ────────────────────────────────────────────────────────
    section("MEDICATION CONFIGURATION", bg=C["text2"])
    col_hdrs()
    mrow("Metformin Start Date",
         f'=IFERROR(TEXT(INT({MED_START_CELL}),"YYYY-MM-DD"),"← Enter date in INPUT cell I1")',
         f'=IFERROR(TEXT(INT({MED_START_CELL}),"YYYY-MM-DD"),"—")',
         note="Date entered in INPUT sheet cell I1",
         num_fmt="@", no_delta=True)
    mrow("Washout Buffer",
         f"={MED_WASHOUT_CELL}", f"={MED_WASHOUT_CELL}",
         unit="days",
         note="Days before & after start date excluded as transition period (INPUT cell I2, default 4)",
         num_fmt="0", no_delta=True)
    mrow("Tagged Readings",
         countif_meds("Off Meds"), countif_meds("On Meds"),
         unit="readings", num_fmt="0",
         note="Rows tagged in _Calc. Washout rows are excluded from all calculations below.")
    blank()

    # ── BASIC STATISTICS ─────────────────────────────────────────────────────
    section("BASIC STATISTICS")
    col_hdrs()
    mrow("Average Glucose",
         avgifs_meds("Off Meds"), avgifs_meds("On Meds"),
         unit="mg/dL",
         note="Mean glucose across each period. Negative Δ = lower average on meds.")
    mrow("Minimum Glucose",
         f'=IFERROR(MINIFS({GLU},{MEDS},"Off Meds"),"—")',
         f'=IFERROR(MINIFS({GLU},{MEDS},"On Meds"),"—")',
         unit="mg/dL", note="Lowest single reading in each period")
    mrow("Maximum Glucose",
         f'=IFERROR(MAXIFS({GLU},{MEDS},"Off Meds"),"—")',
         f'=IFERROR(MAXIFS({GLU},{MEDS},"On Meds"),"—")',
         unit="mg/dL", note="Highest single reading in each period")
    mrow("Hypo Events (<70 mg/dL)",
         f'=IFERROR(COUNTIFS({GLU},"<70",{MEDS},"Off Meds"),0)',
         f'=IFERROR(COUNTIFS({GLU},"<70",{MEDS},"On Meds"),0)',
         unit="readings", num_fmt="0",
         note="Monitor for hypoglycaemia when stopping or starting metformin")
    mrow("Hyperglycemia Burden Score (HBS)",
         hbs_meds("Off Meds"), hbs_meds("On Meds"),
         unit="mg/dL·rdgs",
         note="Sum of excess glucose above 180. Captures frequency AND magnitude — "
              "tighter efficacy signal than TAR% alone. Large negative Δ = significant improvement.")
    blank()

    # ── GLYCEMIC VARIABILITY ─────────────────────────────────────────────────
    section("GLYCEMIC VARIABILITY")
    col_hdrs()
    mrow("Standard Deviation (SD)",
         sd_meds("Off Meds"), sd_meds("On Meds"),
         unit="mg/dL",
         note="Population SD. A drop confirms metformin stabilises glucose, not just lowers the mean.")
    cv_off = f'=IFERROR(B{r-1}/AVERAGEIFS({GLU},{MEDS},"Off Meds")*100,"—")'
    cv_on  = f'=IFERROR(C{r-1}/AVERAGEIFS({GLU},{MEDS},"On Meds")*100,"—")'
    mrow("Coefficient of Variation (CV)",
         cv_off, cv_on, unit="%",
         note="CV = SD ÷ Mean × 100. Target <36% (stable). Tracks if meds reduce erratic swings.")
    mrow("LBGI — Low Blood Glucose Index",
         lbgi_meds("Off Meds"), lbgi_meds("On Meds"),
         unit="score",
         note="Kovatchev low-risk score. <1.1: Low  1.1–2.5: Moderate  >2.5: High. "
              "Rising on meds would flag increased hypo risk.")
    mrow("HBGI — High Blood Glucose Index",
         hbgi_meds("Off Meds"), hbgi_meds("On Meds"),
         unit="score",
         note="Kovatchev high-risk score. <4.5: Low  4.5–9: Moderate  >9: High. "
              "Should fall on metformin (confirms hepatic glucose suppression).")
    blank()

    # ── ESTIMATED HbA1c ───────────────────────────────────────────────────────
    section("ESTIMATED HbA1c")
    col_hdrs()
    mrow("GMI — Glucose Management Indicator",
         gmi_meds("Off Meds"), gmi_meds("On Meds"),
         unit="%", off_color=C["tbr1"], on_color=C["tir"],
         note="3.31 + 0.02392 × mean glucose. ≥7.0%: Diabetic  5.7–6.9%: Pre-diabetic  <5.7%: Normal")
    blank()

    # ── TIME IN RANGE ─────────────────────────────────────────────────────────
    section("TIME IN RANGE  (ADA/EASD Targets — T2D)")
    col_hdrs()
    tir_tiers = [
        ("TBR Level 2  (<54 mg/dL)",
         f'=IFERROR(COUNTIFS({GLU},"<54",{MEDS},"Off Meds")/COUNTIF({MEDS},"Off Meds")*100,0)',
         f'=IFERROR(COUNTIFS({GLU},"<54",{MEDS},"On Meds")/COUNTIF({MEDS},"On Meds")*100,0)',
         "Target <1%"),
        ("TBR Level 1  (54–69 mg/dL)",
         f'=IFERROR(COUNTIFS({GLU},">="&54,{GLU},"<"&70,{MEDS},"Off Meds")/COUNTIF({MEDS},"Off Meds")*100,0)',
         f'=IFERROR(COUNTIFS({GLU},">="&54,{GLU},"<"&70,{MEDS},"On Meds")/COUNTIF({MEDS},"On Meds")*100,0)',
         "Target <4%"),
        ("TIR  (70–180 mg/dL)",
         f'=IFERROR(COUNTIFS({GLU},">="&70,{GLU},"<="&180,{MEDS},"Off Meds")/COUNTIF({MEDS},"Off Meds")*100,0)',
         f'=IFERROR(COUNTIFS({GLU},">="&70,{GLU},"<="&180,{MEDS},"On Meds")/COUNTIF({MEDS},"On Meds")*100,0)',
         "Target >70%"),
        ("TAR Level 1  (181–250 mg/dL)",
         f'=IFERROR(COUNTIFS({GLU},">"&180,{GLU},"<="&250,{MEDS},"Off Meds")/COUNTIF({MEDS},"Off Meds")*100,0)',
         f'=IFERROR(COUNTIFS({GLU},">"&180,{GLU},"<="&250,{MEDS},"On Meds")/COUNTIF({MEDS},"On Meds")*100,0)',
         "Target <25%"),
        ("TAR Level 2  (>250 mg/dL)",
         f'=IFERROR(COUNTIFS({GLU},">250",{MEDS},"Off Meds")/COUNTIF({MEDS},"Off Meds")*100,0)',
         f'=IFERROR(COUNTIFS({GLU},">250",{MEDS},"On Meds")/COUNTIF({MEDS},"On Meds")*100,0)',
         "Target <5%"),
    ]
    for label, f_off, f_on, note in tir_tiers:
        mrow(label, f_off, f_on, unit="%", note=note)
    blank()

    # ── AREA UNDER THE CURVE ──────────────────────────────────────────────────
    section("AREA UNDER THE CURVE  (Trapezoidal — mg/dL·hr)", bg=C["purple"])
    col_hdrs()
    mrow("Total AUC",
         auc_sumif(TRAP_AUC,   "Off Meds"),
         auc_sumif(TRAP_AUC,   "On Meds"),
         unit="mg/dL·hr",
         note="Integral of the glucose-time curve per phase. Gaps >1 hr excluded.")
    mrow("Hyperglycemic AUC  (>180 mg/dL)",
         auc_sumif(TRAP_HYPER, "Off Meds"),
         auc_sumif(TRAP_HYPER, "On Meds"),
         unit="mg/dL·hr", off_color=C["tar1"], on_color=C["tar1"],
         note="Time-weighted exposure above 180 mg/dL. More sensitive than TAR% — "
              "captures both how often AND how high. Large negative Δ confirms metformin efficacy.")
    mrow("Hypoglycemic AUC  (<70 mg/dL)",
         auc_sumif(TRAP_HYPO,  "Off Meds"),
         auc_sumif(TRAP_HYPO,  "On Meds"),
         unit="mg/dL·hr", off_color=C["red"], on_color=C["red"],
         note="Time-weighted deficit below 70 mg/dL. Monitor: an increase on meds "
              "could indicate over-treatment risk if combined with other agents.")
    mrow("Daily Hyper AUC  (per day, normalised)",
         daily_auc_meds(TRAP_HYPER, "Off Meds"),
         daily_auc_meds(TRAP_HYPER, "On Meds"),
         unit="mg/dL·hr/day", off_color=C["tar1"], on_color=C["tar1"],
         note="Hyperglycemic AUC ÷ unique days per phase. Corrects for unequal phase durations — "
              "use this column, not Total AUC, for the primary treatment-effect comparison.")
    mrow("Daily Hypo AUC  (per day, normalised)",
         daily_auc_meds(TRAP_HYPO,  "Off Meds"),
         daily_auc_meds(TRAP_HYPO,  "On Meds"),
         unit="mg/dL·hr/day", off_color=C["red"], on_color=C["red"],
         note="Hypoglycemic AUC ÷ unique days per phase.")
    blank()

    # ── GLUCOSE DISTRIBUTION — PERCENTILES ───────────────────────────────────
    section("GLUCOSE DISTRIBUTION  (Percentile Profile)", bg=C["accent"])
    col_hdrs()
    pctile_rows = [
        ("10th Percentile",  0.10, "Low end of distribution. If this rises on meds, basal suppression is confirmed."),
        ("25th Percentile",  0.25, "Lower IQR boundary."),
        ("Median (50th Pct)", 0.50, "Central tendency. More robust than mean for skewed CGM data."),
        ("75th Percentile",  0.75, "Upper IQR boundary."),
        ("90th Percentile",  0.90, "High end. If this stays unchanged but lower pctiles drop, "
                                   "meds suppress baseline without capping peaks — expected for metformin."),
    ]
    for label, p, note in pctile_rows:
        mrow(label,
             pctile_meds(p, "Off Meds"),
             pctile_meds(p, "On Meds"),
             unit="mg/dL", note=note)
    blank()

    # ── OVERNIGHT & FASTING TIME BLOCKS (Suggestion 3) ───────────────────────
    section("OVERNIGHT & FASTING TIME BLOCKS  (Metformin's Primary Effect Window)", bg=C["amber"])
    col_hdrs(h2="Off Meds Avg", h3="On Meds Avg")
    overnight_blocks = [
        (0,  3,  "12 AM – 3 AM",
         "Overnight hepatic glucose output — metformin's primary suppression window. "
         "Expect the largest Δ here."),
        (3,  6,  "3 AM – 6 AM",
         "Dawn phenomenon — early-morning cortisol/GH-driven rise. Metformin blunts hepatic response."),
        (6,  9,  "6 AM – 9 AM",
         "Fasting / pre-breakfast — the single clearest marker of metformin efficacy. "
         "Target: significant negative Δ."),
        (9,  12, "9 AM – 12 PM",
         "Post-breakfast postprandial — less metformin-sensitive; more diet/insulin-dependent."),
        (12, 15, "12 PM – 3 PM", "Midday postprandial."),
        (15, 18, "3 PM – 6 PM",  "Afternoon / late postprandial."),
        (18, 21, "6 PM – 9 PM",  "Dinner / evening postprandial."),
        (21, 24, "9 PM – 12 AM", "Pre-sleep / evening baseline."),
    ]
    for h_lo, h_hi, label, relevance in overnight_blocks:
        ws.row_dimensions[r].height = 20
        f_off = avgifs_hour_meds(h_lo, h_hi, "Off Meds")
        f_on  = avgifs_hour_meds(h_lo, h_hi, "On Meds")

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = font(bold=True, color=C["text"]); lc.alignment = align("left"); lc.border = border()

        oc = ws.cell(row=r, column=2, value=f_off)
        oc.font = font(bold=True, color=C["tbr1"], size=11)
        oc.alignment = align("center"); oc.border = border(); oc.number_format = "0.0"

        nc = ws.cell(row=r, column=3, value=f_on)
        nc.font = font(bold=True, color=C["tir"], size=11)
        nc.alignment = align("center"); nc.border = border(); nc.number_format = "0.0"

        dv = f'=IFERROR(IF(ISNUMBER(B{r})*ISNUMBER(C{r}),C{r}-B{r},"—"),"—")'
        dc = ws.cell(row=r, column=4, value=dv)
        dc.font = font(bold=True, color=C["purple"], size=11)
        dc.alignment = align("center"); dc.border = border(); dc.number_format = "0.0"

        uc = ws.cell(row=r, column=5, value="mg/dL")
        uc.font = font(color=C["muted"], size=10); uc.alignment = align("center"); uc.border = border()

        rc = ws.cell(row=r, column=6, value=relevance)
        rc.font = font(color=C["muted"], size=10, italic=True)
        rc.alignment = align("left", wrap=True); rc.border = border()
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    interp = ws.cell(row=r, column=1,
        value="Interpretation: If metformin is working, the 12 AM–9 AM blocks should show the "
              "largest negative Δ (hepatic suppression). Minimal change in 9 AM–9 PM blocks is "
              "expected — those are driven by meals and fast-acting insulin, not metformin.")
    interp.font = font(size=9, color=C["accent"], italic=True)
    interp.alignment = align("left", wrap=True)


def build_daily(wb):
    ws = wb.create_sheet("Daily")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.tab_color = C["green"]

    col_w(ws, 1, 16); col_w(ws, 2, 12); col_w(ws, 3, 20)
    col_w(ws, 4, 16); col_w(ws, 5, 16); col_w(ws, 6, 16)
    col_w(ws, 7, 14); col_w(ws, 8, 18); col_w(ws, 9, 16)

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "Daily Glucose Breakdown"
    t.font  = font(bold=True, size=14, color=C["accent"])
    t.alignment = align("left")

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:I2")
    s = ws["A2"]
    s.value = ("Dates auto-populated from your data (Excel 365 / 2021 required). "
               "Rolling 14-day TIR shows when metformin reached steady-state efficacy. "
               "Nocturnal Hypos counts readings <70 mg/dL between midnight and 6 AM.")
    s.font = font(size=9, color=C["muted"], italic=True)
    s.alignment = align("left", wrap=True)

    ws.row_dimensions[3].height = 22
    for col, hdr in enumerate([
        "Date", "Readings", "Avg Glucose", "TIR % (70–180)",
        "TBR % (<70)", "TAR % (>180)", "Status",
        "Rolling 14-Day TIR %", "Nocturnal Hypos"
    ], 1):
        hcell(ws, 3, col, hdr)

    date_formula = (
        f"=IFERROR(SORT(UNIQUE(FILTER({DATE},{DATE}<>\"\"))),"
        "\"No data — paste CGM data in the INPUT sheet\")"
    )
    dc = ws.cell(row=4, column=1, value=date_formula)
    dc.number_format = "YYYY-MM-DD"
    dc.font = font(color=C["text2"]); dc.alignment = align("center"); dc.border = border()

    for r in range(4, 370):
        date_ref = f"A{r}"

        dcell(ws, r, 2,
              f'=IFERROR(IF({date_ref}="","",COUNTIFS({DATE},{date_ref},{GLU},"<>")),"")',
              num_fmt="0")

        dcell(ws, r, 3,
              f'=IFERROR(IF({date_ref}="","",AVERAGEIFS({GLU},{DATE},{date_ref})),"")',
              num_fmt="0.0")

        dcell(ws, r, 4,
              f'=IFERROR(IF({date_ref}="","",COUNTIFS({GLU},">="&70,{GLU},"<="&180,{DATE},{date_ref})'
              f'/COUNTIFS({DATE},{date_ref},{GLU},"<>")*100),"")',
              num_fmt='0.0"%"')

        dcell(ws, r, 5,
              f'=IFERROR(IF({date_ref}="","",COUNTIFS({GLU},"<70",{DATE},{date_ref})'
              f'/COUNTIFS({DATE},{date_ref},{GLU},"<>")*100),"")',
              num_fmt='0.0"%"')

        dcell(ws, r, 6,
              f'=IFERROR(IF({date_ref}="","",COUNTIFS({GLU},">180",{DATE},{date_ref})'
              f'/COUNTIFS({DATE},{date_ref},{GLU},"<>")*100),"")',
              num_fmt='0.0"%"')

        dcell(ws, r, 7,
              f'=IFERROR(IF({date_ref}="","",IF(D{r}>=70,"✓ Good",IF(E{r}>4,"Low glucose","High glucose"))),"")' ,
              num_fmt="@")

        # Rolling 14-day TIR: readings in range over the trailing 14-day window
        dcell(ws, r, 8,
              f'=IFERROR(IF({date_ref}="","",COUNTIFS({GLU},">="&70,{GLU},"<="&180,'
              f'{DATE},">="&({date_ref}-13),{DATE},"<="&{date_ref})'
              f'/COUNTIFS({DATE},">="&({date_ref}-13),{DATE},"<="&{date_ref},{GLU},"<>")*100),"")',
              num_fmt='0.0"%"')

        # Nocturnal hypos: readings <70 between midnight and 5:59 AM
        dcell(ws, r, 9,
              f'=IFERROR(IF({date_ref}="","",COUNTIFS({GLU},"<70",{DATE},{date_ref},{HOUR},"<6")),"")',
              num_fmt="0")

    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill as PF

    ws.conditional_formatting.add("D4:D369",
        CellIsRule(operator="greaterThanOrEqual", formula=["70"],
                   fill=PF("solid", fgColor="ECFDF5"),
                   font=Font(color=C["green"], bold=True, name="Calibri")))
    ws.conditional_formatting.add("D4:D369",
        CellIsRule(operator="lessThan", formula=["50"],
                   fill=PF("solid", fgColor="FEF2F2"),
                   font=Font(color=C["red"], bold=True, name="Calibri")))
    # Rolling TIR: same colour rules
    ws.conditional_formatting.add("H4:H369",
        CellIsRule(operator="greaterThanOrEqual", formula=["70"],
                   fill=PF("solid", fgColor="ECFDF5"),
                   font=Font(color=C["green"], bold=True, name="Calibri")))
    ws.conditional_formatting.add("H4:H369",
        CellIsRule(operator="lessThan", formula=["50"],
                   fill=PF("solid", fgColor="FEF2F2"),
                   font=Font(color=C["red"], bold=True, name="Calibri")))
    # Nocturnal hypos: flag any non-zero day in red
    ws.conditional_formatting.add("I4:I369",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PF("solid", fgColor="FEF2F2"),
                   font=Font(color=C["red"], bold=True, name="Calibri")))


def build_agp(wb):
    ws = wb.create_sheet("AGP")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.tab_color = C["purple"]

    col_w(ws, 1, 12); col_w(ws, 2, 10)
    for c in range(3, 9): col_w(ws, c, 18)

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Ambulatory Glucose Profile (AGP) — Hourly Percentiles"
    t.font  = font(bold=True, size=14, color=C["accent"])
    t.alignment = align("left")

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = ("Median glucose and interquartile range (25th–75th pct) by hour of day. "
               "Target range 70–180 mg/dL shown for reference. Requires Excel 365.")
    s.font = font(size=9, color=C["muted"], italic=True)
    s.alignment = align("left")

    ws.row_dimensions[3].height = 22
    for col, hdr in enumerate(["Hour", "Readings", "10th pct", "25th pct (IQR low)",
                                "Median (50th)", "75th pct (IQR high)", "90th pct", "Zone"], 1):
        hcell(ws, 3, col, hdr)

    for i, h in enumerate(range(24)):
        r = i + 4
        dcell(ws, r, 1, f"{h:02d}:00", bold=True, color=C["accent"])
        dcell(ws, r, 2, f"=IFERROR(COUNTIF({HOUR},{h}),0)", num_fmt="0")
        dcell(ws, r, 3, pctile_hour(h, 0.10), num_fmt="0.0")
        dcell(ws, r, 4, pctile_hour(h, 0.25), num_fmt="0.0")
        dcell(ws, r, 5, pctile_hour(h, 0.50), num_fmt="0.0", bold=True, color=C["accent"])
        dcell(ws, r, 6, pctile_hour(h, 0.75), num_fmt="0.0")
        dcell(ws, r, 7, pctile_hour(h, 0.90), num_fmt="0.0")
        zone_formula = (
            f'=IFERROR(IF({pctile_hour(h,0.50)[1:]}="","—",'
            f'IF({pctile_hour(h,0.50)[1:]}<70,"Low",'
            f'IF({pctile_hour(h,0.50)[1:]}>180,"High","In Range"))),"—")'
        )
        dcell(ws, r, 8, zone_formula, num_fmt="@")

    r2 = 28
    ws.merge_cells(f"A{r2}:H{r2}")
    note = ws.cell(row=r2, column=1,
                   value="Reference: Target range 70–180 mg/dL  |  IQR band = 25th–75th percentile  |  Solid line = median")
    note.font = font(size=9, color=C["muted"], italic=True)
    note.alignment = align("left")


def build_timeblocks(wb):
    ws = wb.create_sheet("Time Blocks")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["amber"]

    col_w(ws, 1, 16); col_w(ws, 2, 12); col_w(ws, 3, 18)
    col_w(ws, 4, 16); col_w(ws, 5, 16); col_w(ws, 6, 14)
    col_w(ws, 7, 12); col_w(ws, 8, 40)

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Average Glucose by 3-Hour Time Block"
    t.font  = font(bold=True, size=14, color=C["accent"])
    t.alignment = align("left")

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = ("Average glucose per 3-hour window — identifies postprandial hyperglycaemia patterns "
               "and dawn phenomenon. Off Meds / On Meds split requires a start date in INPUT cell I1.")
    s.font = font(size=9, color=C["muted"], italic=True)
    s.alignment = align("left")

    ws.row_dimensions[3].height = 22
    for col, hdr in enumerate([
        "Time Window", "Readings", "Overall Avg",
        "Off Meds Avg", "On Meds Avg", "Δ Change",
        "Zone", "Clinical Relevance"
    ], 1):
        hcell(ws, 3, col, hdr)

    blocks = [
        (0,  3,  "12–3 AM",  "Overnight baseline — hepatic glucose output"),
        (3,  6,  "3–6 AM",   "Dawn phenomenon window — early-morning glucose rise"),
        (6,  9,  "6–9 AM",   "Fasting / pre-breakfast — key metformin efficacy marker"),
        (9,  12, "9–12 PM",  "Post-breakfast — postprandial excursion"),
        (12, 15, "12–3 PM",  "Lunch / midday postprandial"),
        (15, 18, "3–6 PM",   "Afternoon — late postprandial / activity"),
        (18, 21, "6–9 PM",   "Dinner / evening postprandial"),
        (21, 24, "9–12 AM",  "Evening / pre-sleep"),
    ]

    for i, (h_lo, h_hi, label, relevance) in enumerate(blocks, 4):
        dcell(ws, i, 1, label, bold=True, h_align="left")
        dcell(ws, i, 2,
              f"=IFERROR(COUNTIFS({HOUR},\">=\"&{h_lo},{HOUR},\"<\"&{h_hi}),0)",
              num_fmt="0")
        avg_formula = avgifs_hour(h_lo, h_hi)
        dcell(ws, i, 3, avg_formula, num_fmt="0.0", bold=True)
        dcell(ws, i, 4, avgifs_hour_meds(h_lo, h_hi, "Off Meds"),
              num_fmt="0.0", color=C["tbr1"])
        dcell(ws, i, 5, avgifs_hour_meds(h_lo, h_hi, "On Meds"),
              num_fmt="0.0", color=C["tir"])
        delta_f = f'=IFERROR(IF(ISNUMBER(D{i})*ISNUMBER(E{i}),E{i}-D{i},"—"),"—")'
        dcell(ws, i, 6, delta_f, num_fmt="0.0", color=C["purple"])
        zone_formula = (
            f'=IFERROR(IF({avg_formula[1:]}="","—",'
            f'IF({avg_formula[1:]}<70,"Low",'
            f'IF({avg_formula[1:]}>180,"High","In Range"))),"—")'
        )
        dcell(ws, i, 7, zone_formula, num_fmt="@")
        rc = ws.cell(row=i, column=8, value=relevance)
        rc.font = font(color=C["muted"], size=10, italic=True)
        rc.alignment = align("left"); rc.border = border()

    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill as PF
    for col_range in ["C4:C11", "D4:D11", "E4:E11"]:
        ws.conditional_formatting.add(col_range,
            CellIsRule(operator="greaterThan", formula=["180"],
                       fill=PF("solid", fgColor="FFFBEB"),
                       font=Font(color=C["amber"], bold=True, name="Calibri")))
        ws.conditional_formatting.add(col_range,
            CellIsRule(operator="lessThan", formula=["70"],
                       fill=PF("solid", fgColor="FEF2F2"),
                       font=Font(color=C["red"], bold=True, name="Calibri")))
        ws.conditional_formatting.add(col_range,
            CellIsRule(operator="between", formula=["70", "180"],
                       fill=PF("solid", fgColor="ECFDF5"),
                       font=Font(color=C["green"], bold=True, name="Calibri")))


def build_glossary(wb):
    """Reference glossary of all clinical and statistical terms used in this workbook."""
    ws = wb.create_sheet("Glossary")
    ws.sheet_view.showGridLines = False
    ws.tab_color = C["text2"]

    col_w(ws, 1, 20); col_w(ws, 2, 30); col_w(ws, 3, 62); col_w(ws, 4, 28)

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "LingoCGM — Clinical & Statistical Glossary"
    t.font  = font(bold=True, size=16, color="FFFFFF")
    t.fill  = fill(C["text2"])
    t.alignment = align("left")

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value = ("Definitions for every metric, index, and clinical term used in this workbook. "
               "All ADA/EASD targets reflect the 2023 Consensus Guidelines for adults with T2D.")
    s.font = font(size=9, color=C["muted"], italic=True)
    s.alignment = align("left")

    r = 4

    def section(title, bg=None):
        nonlocal r
        ws.row_dimensions[r].height = 22
        hcell(ws, r, 1, title, bg=bg or C["accent"], size=11, h_align="left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        ws.row_dimensions[r].height = 18
        for col, hdr in enumerate(["Abbreviation", "Full Name", "Definition", "Target / Range"], 1):
            hcell(ws, r, col, hdr, bg=C["text2"], size=10)
        r += 1

    def term(abbr, full, defn, target=""):
        nonlocal r
        ws.row_dimensions[r].height = 38
        bg_row = "FFFFFF" if r % 2 == 0 else C["hdr_alt"]

        ac = ws.cell(row=r, column=1, value=abbr)
        ac.font = font(bold=True, color=C["accent"], size=10)
        ac.alignment = align("left"); ac.border = border(); ac.fill = fill(bg_row)

        fc = ws.cell(row=r, column=2, value=full)
        fc.font = font(bold=False, color=C["text2"], size=10)
        fc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        fc.border = border(); fc.fill = fill(bg_row)

        dc = ws.cell(row=r, column=3, value=defn)
        dc.font = font(bold=False, color=C["text"], size=10)
        dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dc.border = border(); dc.fill = fill(bg_row)

        tc = ws.cell(row=r, column=4, value=target)
        tc.font = font(bold=bool(target), color=C["green"] if target else C["muted"], size=10)
        tc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        tc.border = border(); tc.fill = fill(bg_row)
        r += 1

    # ── CGM DEVICE & DATA ─────────────────────────────────────────────────────
    section("CGM DEVICE & DATA", bg=C["accent"])
    term("CGM",
         "Continuous Glucose Monitor",
         "A small sensor (usually on the arm or abdomen) that measures interstitial glucose "
         "every 5–15 minutes and transmits readings wirelessly. Interstitial glucose lags blood "
         "glucose by ~5–10 minutes.",
         "Wear time ≥70%")
    term("DateTime",
         "Timestamp of Each Reading",
         "The exact date and time each glucose reading was recorded. Required in column A of the "
         "INPUT sheet in YYYY-MM-DD HH:MM:SS format.",
         "")
    term("mg/dL",
         "Milligrams per Decilitre",
         "The unit used in the United States for blood glucose concentration. To convert to mmol/L "
         "(used in the UK, Canada, Australia): divide by 18.0182.",
         "Normal fasting: 70–100 mg/dL")

    # ── CORE TIME-IN-RANGE METRICS ────────────────────────────────────────────
    section("CORE TIME-IN-RANGE METRICS  (ADA/EASD 2023)", bg=C["tir"])
    term("TIR",
         "Time in Range  (70–180 mg/dL)",
         "The percentage of CGM readings within the target glucose range of 70–180 mg/dL. "
         "The primary CGM outcome metric endorsed by the ADA and EASD. Each 5% increase in TIR "
         "is associated with clinically meaningful reductions in HbA1c and complication risk.",
         "T2D target: >70%")
    term("TBR L1",
         "Time Below Range Level 1  (54–69 mg/dL)",
         "Percentage of readings between 54–69 mg/dL. Represents clinically significant "
         "hypoglycaemia. Values above target warrant review of insulin dose or medication.",
         "Target: <4%")
    term("TBR L2",
         "Time Below Range Level 2  (<54 mg/dL)",
         "Percentage of readings below 54 mg/dL. Represents serious hypoglycaemia. "
         "Even 1% is clinically concerning and warrants immediate medication review.",
         "Target: <1%")
    term("TAR L1",
         "Time Above Range Level 1  (181–250 mg/dL)",
         "Percentage of readings between 181–250 mg/dL. Moderate hyperglycaemia. "
         "Sustained elevation contributes to glycation and microvascular damage.",
         "Target: <25%")
    term("TAR L2",
         "Time Above Range Level 2  (>250 mg/dL)",
         "Percentage of readings above 250 mg/dL. Significant hyperglycaemia. "
         "Values above target indicate poor glycaemic control.",
         "Target: <5%")

    # ── ESTIMATED HbA1c ───────────────────────────────────────────────────────
    section("ESTIMATED HbA1c", bg=C["accent"])
    term("GMI",
         "Glucose Management Indicator",
         "ADA/EASD preferred estimate of HbA1c from CGM data. Formula: 3.31 + 0.02392 × mean "
         "glucose (mg/dL). More accurate than eA1c for CGM datasets. Does not replace laboratory "
         "HbA1c for clinical decisions.",
         "T2D target: <7.0%")
    term("eA1c",
         "Estimated A1c (Legacy Formula)",
         "Older estimate: (mean glucose + 46.7) ÷ 28.7. Less accurate than GMI for CGM data "
         "because it was derived from fingerstick measurements. Included for reference only.",
         "T2D target: <7.0%")

    # ── GLYCEMIC VARIABILITY ──────────────────────────────────────────────────
    section("GLYCEMIC VARIABILITY", bg=C["purple"])
    term("SD",
         "Standard Deviation",
         "Statistical spread of all glucose readings around the mean. A lower SD means glucose "
         "stays closer to the average. Population SD (STDEV.P) is used here, which treats the "
         "dataset as the complete population rather than a sample.",
         "T2D: aim <36 mg/dL")
    term("CV",
         "Coefficient of Variation",
         "SD ÷ Mean × 100. The gold-standard metric for glycaemic variability because it is "
         "independent of the mean glucose level, making it valid for comparing periods with "
         "different average glucose (e.g., Off Meds vs On Meds).",
         "Stable: <36%  High: >50%")
    term("MAGE",
         "Mean Amplitude of Glycemic Excursions",
         "The average magnitude of glucose swings that exceed one standard deviation. Captures "
         "the clinical impact of peaks and troughs. Cannot be computed with Excel array formulas "
         "— requires the Python script (cgm_to_xls.py) which implements the Kovatchev algorithm.",
         "Lower is better")
    term("IQR",
         "Interquartile Range  (25th–75th Pct)",
         "The band between the 25th and 75th percentile of glucose values. In the AGP chart, "
         "this band represents the 'typical' daily glucose range — 50% of readings fall within it. "
         "A narrow IQR indicates consistent daily patterns.",
         "")

    # ── KOVATCHEV RISK INDICES ────────────────────────────────────────────────
    section("KOVATCHEV RISK INDICES  (FDA-Validated)", bg=C["purple"])
    term("LBGI",
         "Low Blood Glucose Index",
         "FDA-validated risk score for hypoglycaemia. Computed as the mean of 10 × f(BG)² for "
         "all readings where f(BG) < 0. Higher values indicate greater hypo risk. Developed by "
         "Dr. Boris Kovatchev at UVA and used in artificial pancreas research.",
         "<1.1: Low  1.1–2.5: Moderate  >2.5: High")
    term("HBGI",
         "High Blood Glucose Index",
         "FDA-validated risk score for hyperglycaemia. Computed as the mean of 10 × f(BG)² for "
         "all readings where f(BG) > 0. Should decrease on metformin as it suppresses hepatic "
         "glucose output. Developed alongside LBGI by Kovatchev et al.",
         "<4.5: Low  4.5–9: Moderate  >9: High")
    term("f(BG)",
         "Kovatchev Risk Transform",
         "The symmetric transformation applied to each glucose value: "
         "f(BG) = 1.509 × (ln(BG)^1.084 − 5.381). Negative values indicate hypoglycaemic risk; "
         "positive values indicate hyperglycaemic risk. The squaring step (×10 × f²) penalises "
         "extreme values proportionally.",
         "f = 0 at ~112 mg/dL")
    term("HBS",
         "Hyperglycemia Burden Score",
         "The sum of (glucose − 180) across all readings above 180 mg/dL. Unlike TAR%, which "
         "only counts frequency, HBS captures both frequency AND magnitude: two long readings "
         "at 350 mg/dL score much higher than six brief readings at 185 mg/dL. "
         "Preferred treatment-effect signal when comparing medication periods.",
         "Lower is better; target 0")

    # ── AGP & PATTERNS ────────────────────────────────────────────────────────
    section("AMBULATORY GLUCOSE PROFILE & PATTERNS", bg=C["accent"])
    term("AGP",
         "Ambulatory Glucose Profile",
         "A standardised visualisation showing glucose percentiles (10th, 25th, 50th, 75th, "
         "90th) plotted by hour of day across the entire study period. The AGP compresses weeks "
         "of data into a single representative 24-hour curve. Endorsed by the ADA as the "
         "standard CGM report format.",
         "")
    term("Dawn Phenomenon",
         "Early-Morning Glucose Rise",
         "A natural rise in blood glucose between approximately 3 AM and 8 AM, driven by "
         "overnight surges of cortisol and growth hormone that stimulate the liver to release "
         "stored glucose. Metformin's primary mechanism (suppressing hepatic glucose output) "
         "directly blunts the dawn phenomenon.",
         "Visible in 3–6 AM time block")
    term("Nocturnal Hypo",
         "Nocturnal Hypoglycaemia",
         "A reading below 70 mg/dL occurring between midnight and 6 AM. Clinically important "
         "because the patient is asleep and unlikely to notice symptoms. The Daily sheet flags "
         "any night with at least one nocturnal hypo reading.",
         "Target: 0 events/night")
    term("Fasting Glucose",
         "Pre-Breakfast Fasting Glucose",
         "Glucose measured after ≥8 hours without eating. Approximated in this workbook by the "
         "6–9 AM time block. This is the single clearest marker of metformin efficacy because "
         "it directly reflects overnight hepatic glucose production.",
         "Normal: 70–100 mg/dL")
    term("Postprandial",
         "Post-Meal Glucose Excursion",
         "The rise in glucose following a meal, typically peaking 60–90 minutes after eating. "
         "Captured by the 9 AM–12 PM, 12–3 PM, and 6–9 PM time blocks. Metformin has limited "
         "direct effect on postprandial spikes — those respond better to diet or rapid-acting insulin.",
         "Peak <180 mg/dL")

    # ── MEDICATION & STUDY DESIGN ─────────────────────────────────────────────
    section("MEDICATION & STUDY DESIGN", bg=C["text2"])
    term("Metformin",
         "Biguanide Oral Hypoglycaemic Agent",
         "First-line oral medication for Type 2 Diabetes. Primary mechanism: suppresses hepatic "
         "glucose production (gluconeogenesis). Secondary effects: improves peripheral insulin "
         "sensitivity and reduces intestinal glucose absorption. Does not cause hypoglycaemia "
         "when used alone. Reaches steady-state plasma concentration in 24–48 hours.",
         "")
    term("Washout",
         "Pharmacological Washout Period",
         "Days excluded from analysis around the medication start date to avoid contaminating "
         "'clean' Off Meds or On Meds periods with transition-state data. For metformin, "
         "3–5 days on each side accounts for the time to reach (or clear) steady-state "
         "tissue levels. Set in INPUT cell I2 (default: 4 days).",
         "Recommended: 3–5 days")
    term("Off Meds",
         "Pre-Medication Baseline Period",
         "All CGM data recorded before the washout window. Represents the patient's natural "
         "glucose pattern without metformin. Used as the baseline comparator in all "
         "Comparative sheet calculations.",
         "")
    term("On Meds",
         "Post-Washout Treatment Period",
         "All CGM data recorded after the washout window. Represents the patient's glucose "
         "pattern at metformin steady-state. The Comparative sheet calculates Δ = On Meds − Off Meds "
         "for every metric; a negative Δ is generally favourable.",
         "")
    term("Hepatic Glucose",
         "Hepatic Glucose Output (HGO)",
         "The rate at which the liver releases glucose into the bloodstream from stored glycogen "
         "(glycogenolysis) and newly synthesised glucose (gluconeogenesis). HGO is the dominant "
         "driver of fasting and nocturnal glucose in T2D. Metformin's primary action is to reduce HGO "
         "by activating AMPK in hepatocytes.",
         "")

    # ── AREA UNDER THE CURVE ─────────────────────────────────────────────────
    section("AREA UNDER THE CURVE  (AUC)", bg=C["purple"])
    term("AUC",
         "Area Under the Glucose-Time Curve",
         "The integral of the glucose concentration over time, computed using the trapezoidal "
         "rule: each consecutive pair of readings contributes ½ × (g1 + g2) × Δt hours. "
         "Provides a single number representing total glucose exposure. Intervals with gaps "
         ">1 hour between readings are excluded to avoid artefacts from sensor downtime.",
         "Units: mg/dL·hr")
    term("Hyper AUC",
         "Hyperglycemic AUC  (>180 mg/dL)",
         "The area of the glucose-time curve that lies above the 180 mg/dL threshold. "
         "Formula per interval: ½ × (max(0,g1−180) + max(0,g2−180)) × Δt. "
         "Captures both the frequency and depth of hyperglycaemia in a single number. "
         "Preferred over TAR% when comparing treatment periods because TAR% is insensitive "
         "to the severity of excursions — two readings at 350 mg/dL score the same as two "
         "readings at 185 mg/dL in TAR%, but not in Hyper AUC.",
         "Lower is better")
    term("Hypo AUC",
         "Hypoglycemic AUC  (<70 mg/dL)",
         "The area of the glucose-time curve that lies below the 70 mg/dL threshold. "
         "Formula per interval: ½ × (max(0,70−g1) + max(0,70−g2)) × Δt. "
         "Reflects both how long and how deep hypoglycaemia episodes are.",
         "Lower is better; target 0")
    term("Daily AUC",
         "AUC Normalised per Calendar Day",
         "Total AUC for a phase divided by the number of unique calendar days in that phase. "
         "Critical for comparing Off Meds vs On Meds periods that have different durations — "
         "the Off Meds period is often longer (full baseline) than the On Meds period "
         "(limited monitoring window), so raw totals are not directly comparable.",
         "Use for primary comparison")
    term("Trapezoidal Rule",
         "Numerical Integration Method",
         "The standard algorithm for computing AUC from discrete time-series data. "
         "Each pair of consecutive readings (t1,g1) and (t2,g2) contributes a trapezoid "
         "of area = (g1+g2)/2 × (t2−t1). More accurate than the rectangle method for "
         "CGM data because it accounts for the linear interpolation between readings.",
         "")
    term("Gap Exclusion",
         "Sensor Gap Threshold (>1 hour)",
         "Intervals between consecutive readings longer than 1 hour are excluded from all "
         "AUC calculations. This prevents large artefactual AUC contributions from periods "
         "when the sensor was off, charging, or experiencing signal loss. "
         "The 1-hour threshold works for both 5-minute and 15-minute CGM devices.",
         "")

    # ── STATISTICAL METHODS ───────────────────────────────────────────────────
    section("STATISTICAL METHODS USED IN THIS WORKBOOK", bg=C["accent"])
    term("Rolling 14-Day TIR",
         "14-Day Trailing Time in Range",
         "TIR recalculated each day using only the 14 most recent calendar days of data. "
         "In the Daily sheet, this rolling window shows when metformin's effect stabilised — "
         "the curve typically rises over the first 7–14 days then plateaus at the new steady-state TIR.",
         "")
    term("Percentile",
         "Percentile Distribution",
         "The value below which a given percentage of observations fall. In the Comparative sheet, "
         "the 10th/25th/50th/75th/90th percentiles show whether metformin shifts the entire glucose "
         "distribution downward or only affects specific parts. Metformin typically lowers the lower "
         "half (baseline) while post-meal peaks remain similar.",
         "")
    term("AVERAGEIFS",
         "Conditional Average (Excel)",
         "Excel function that computes the mean of values meeting one or more criteria. Used "
         "throughout this workbook to calculate phase-conditioned metrics (e.g., average glucose "
         "where Med_Status = 'On Meds'). Requires Excel 2007 or later.",
         "")
    term("FILTER + PERCENTILE",
         "Conditional Percentile (Excel 365)",
         "No native PERCENTILEIFS function exists in Excel. This workbook uses "
         "PERCENTILE(FILTER(range, condition), p) to compute phase-conditioned percentiles. "
         "Requires Excel 365 or Excel 2021.",
         "")
    term("Cond. SD (SUMPRODUCT)",
         "Conditional Standard Deviation Workaround",
         "Excel has no STDEVIFS function. This workbook computes conditional population SD using: "
         "SQRT(SUMPRODUCT((phase_match) × ISNUMBER(glucose) × (glucose − phase_mean)²) ÷ count). "
         "Mathematically equivalent to STDEV.P applied to the filtered subset.",
         "")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    out = "LingoCGM_CGM_Report.xlsx"
    print(f"Building {out}  (NROWS={NROWS:,})…")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("  INPUT sheet…")
    build_input(wb)
    print("  _Calc sheet…")
    build_calc(wb)
    print("  Summary sheet…")
    build_summary(wb)
    print("  Comparative sheet…")
    build_comparative_summary(wb)
    print("  Daily sheet…")
    build_daily(wb)
    print("  AGP sheet…")
    build_agp(wb)
    print("  Time Blocks sheet…")
    build_timeblocks(wb)
    print("  Glossary sheet…")
    build_glossary(wb)

    # Logical clinical flow: data entry → overview → treatment analysis →
    # daily trend → hourly profile → time windows → reference → hidden calc
    sheet_order = [
        "INPUT", "Summary", "Comparative",
        "Daily", "AGP", "Time Blocks",
        "Glossary", "_Calc",
    ]
    wb._sheets.sort(key=lambda s: sheet_order.index(s.title)
                    if s.title in sheet_order else 99)

    print("  Saving…", end="", flush=True)
    wb.save(out)
    print(f" done → {out}")

    import os
    size_kb = os.path.getsize(out) / 1024
    print(f"  File size: {size_kb:.0f} KB")
    print(f"  Data capacity: {NROWS:,} rows  "
          f"(≈ {NROWS//288} days at 5-min  /  ≈ {NROWS//96} days at 15-min)")
    print("\n  Sheet order (clinical flow):")
    for s in wb.worksheets:
        state = " [hidden]" if s.sheet_state == "hidden" else ""
        print(f"    {s.title}{state}")
    print("\n  ✓  Open LingoCGM_CGM_Report.xlsx in Excel 365, paste CGM data")
    print("     in the INPUT sheet (col A = DateTime, col B = Glucose mg/dL).")
    print("     Enter your metformin start date in INPUT cell I1.")
    print("     All other sheets recalculate automatically.")
    print("     See the Glossary sheet for definitions of every metric.")


if __name__ == "__main__":
    main()
